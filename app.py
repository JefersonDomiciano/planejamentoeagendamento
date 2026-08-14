import datetime
import io
import html
import requests
import pandas as pd
import streamlit as st
import openpyxl

from fpdf import FPDF
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================
# CONFIGURAÇÃO
# ============================================================

st.set_page_config(
    page_title="Gestão de Cargas - Logística",
    page_icon="🚚",
    layout="wide",
    initial_sidebar_state="collapsed",
)

FIREBASE_PROJECT_ID = "logistica-d6c14"

FIREBASE_BASE = (
    f"https://firestore.googleapis.com/v1/projects/"
    f"{FIREBASE_PROJECT_ID}/databases/(default)/documents"
)


# ============================================================
# STATUS DO KANBAN
# ============================================================

STATUS = [
    "Aguardando Carregamento",
    "Carregado / No Pátio",
    "Em Trânsito / Viagem Iniciada",
    "Entregue / Concluído",
]

STATUS_ICONS = {
    "Aguardando Carregamento": "⏳",
    "Carregado / No Pátio": "📦",
    "Em Trânsito / Viagem Iniciada": "🚚",
    "Entregue / Concluído": "✅",
}

STATUS_COLORS = {
    "Aguardando Carregamento": "#3b82f6",
    "Carregado / No Pátio": "#22c55e",
    "Em Trânsito / Viagem Iniciada": "#f59e0b",
    "Entregue / Concluído": "#a855f7",
}


# ============================================================
# CSS
# ============================================================

st.markdown(
    """
    <style>

        #MainMenu, footer {
            visibility:hidden;
        }

        .block-container {
            padding-top:1.1rem;
            padding-bottom:2rem;
            max-width:1550px;
        }

        .app-kicker {
            color:#8ea0b8;
            font-size:11px;
            font-weight:800;
            text-transform:uppercase;
            letter-spacing:1.5px;
        }

        .app-title {
            color:#f8fafc;
            font-size:30px;
            font-weight:850;
            letter-spacing:-.6px;
            margin:0;
        }

        .app-subtitle {
            color:#8796aa;
            font-size:13px;
            margin-top:3px;
        }

        /* =========================
           MÉTRICAS
        ========================= */

        .metric-card {
            position:relative;
            overflow:hidden;
            background:linear-gradient(
                145deg,
                #172131 0%,
                #111827 100%
            );
            border:1px solid #263449;
            border-radius:14px;
            padding:16px 18px;
            min-height:112px;
            box-shadow:0 10px 28px rgba(0,0,0,.14);
        }

        .metric-title {
            color:#8998ac;
            font-size:10px;
            font-weight:800;
            letter-spacing:.8px;
        }

        .metric-value {
            color:#f8fafc;
            font-size:29px;
            font-weight:850;
            line-height:1.1;
            margin-top:7px;
        }

        .metric-subtitle {
            color:#718198;
            font-size:10px;
            margin-top:6px;
        }

        .metric-blue {
            border-left:3px solid #3b82f6;
        }

        .metric-green {
            border-left:3px solid #22c55e;
        }

        .metric-yellow {
            border-left:3px solid #f59e0b;
        }

        .metric-purple {
            border-left:3px solid #a855f7;
        }

        .metric-red {
            border-left:3px solid #ef4444;
        }


        /* =========================
           ALERTAS
        ========================= */

        .alert-box {
            padding:12px 15px;
            border-radius:11px;
            margin:8px 0;
            font-size:12px;
            font-weight:650;
        }

        .alert-red {
            background:rgba(239,68,68,.08);
            border:1px solid rgba(239,68,68,.28);
            color:#fca5a5;
        }

        .alert-yellow {
            background:rgba(245,158,11,.08);
            border:1px solid rgba(245,158,11,.28);
            color:#fcd34d;
        }

        .alert-green {
            background:rgba(34,197,94,.08);
            border:1px solid rgba(34,197,94,.25);
            color:#86efac;
        }


        /* =========================
           KANBAN
        ========================= */

        .kanban-column {
            background:rgba(15,23,42,.35);
            border:1px solid #263449;
            border-radius:14px;
            padding:10px;
            min-height:300px;
        }

        .kanban-header {
            background:linear-gradient(
                135deg,
                #182235 0%,
                #111827 100%
            );
            color:#f8fafc;
            padding:12px 13px;
            border-radius:12px;
            font-weight:750;
            font-size:12px;
            border:1px solid #273449;
            margin-bottom:10px;
        }

        .kanban-count {
            color:#71819a;
            font-size:10px;
            font-weight:600;
            margin-top:3px;
        }

        .kanban-empty {
            padding:30px 10px;
            text-align:center;
            color:#536277;
            font-size:11px;
        }

        .kanban-card {
            background:linear-gradient(
                145deg,
                #172131 0%,
                #111827 100%
            );
            border:1px solid #263449;
            border-radius:14px;
            padding:14px;
            margin-bottom:8px;
            box-shadow:0 10px 26px rgba(0,0,0,.14);
        }

        .card-label {
            color:#73849a;
            font-size:10px;
            font-weight:700;
            text-transform:uppercase;
            letter-spacing:.7px;
        }

        .card-id {
            color:#60a5fa;
            font-size:12px;
            font-weight:800;
            margin-bottom:4px;
        }

        .card-driver {
            color:#f8fafc;
            font-size:14px;
            font-weight:800;
            margin:5px 0 8px;
        }

        .card-destination {
            color:#d6deea;
            font-size:12px;
            font-weight:650;
        }

        .card-meta {
            color:#a5b3c5;
            font-size:10px;
            line-height:1.65;
        }

        .card-divider {
            height:1px;
            background:#243145;
            margin:9px 0;
        }

        .card-status {
            display:inline-block;
            padding:5px 8px;
            border-radius:999px;
            font-size:8px;
            font-weight:800;
            margin-bottom:8px;
        }

        .badge {
            display:inline-block;
            padding:4px 7px;
            border-radius:999px;
            font-size:8px;
            font-weight:800;
            letter-spacing:.25px;
            margin-right:3px;
            margin-bottom:5px;
        }

        .badge-red {
            background:rgba(239,68,68,.12);
            color:#fca5a5;
            border:1px solid rgba(239,68,68,.24);
        }

        .badge-yellow {
            background:rgba(245,158,11,.12);
            color:#fcd34d;
            border:1px solid rgba(245,158,11,.24);
        }

        .badge-green {
            background:rgba(34,197,94,.12);
            color:#86efac;
            border:1px solid rgba(34,197,94,.22);
        }

        .badge-blue {
            background:rgba(59,130,246,.12);
            color:#93c5fd;
            border:1px solid rgba(59,130,246,.22);
        }

        .status-help {
            color:#718198;
            font-size:9px;
            margin-top:4px;
            margin-bottom:8px;
        }


        /* =========================
           SEÇÕES / GRÁFICOS
        ========================= */

        .section-title {
            color:#f1f5f9;
            font-size:18px;
            font-weight:800;
            margin:6px 0 13px;
        }

        .chart-card {
            background:linear-gradient(
                145deg,
                #172131 0%,
                #111827 100%
            );
            border:1px solid #263449;
            border-radius:14px;
            padding:18px;
            margin-bottom:14px;
        }

        .chart-heading {
            color:#f1f5f9;
            font-size:14px;
            font-weight:800;
            margin-bottom:14px;
        }

        .chart-heading small {
            color:#718198;
            float:right;
            font-size:10px;
            font-weight:600;
        }

        .bar-list {
            display:flex;
            flex-direction:column;
            gap:12px;
        }

        .bar-row {
            display:grid;
            grid-template-columns:190px 1fr 35px;
            gap:10px;
            align-items:center;
        }

        .bar-label {
            overflow:hidden;
            text-overflow:ellipsis;
            white-space:nowrap;
            color:#cbd5e1;
            font-size:11px;
        }

        .bar-track {
            height:9px;
            background:#202c3d;
            border-radius:999px;
            overflow:hidden;
        }

        .bar-fill {
            height:100%;
            border-radius:999px;
            background:linear-gradient(
                90deg,
                #3b82f6,
                #60a5fa
            );
        }

        .bar-value {
            text-align:right;
            color:#f8fafc;
            font-size:12px;
            font-weight:800;
        }

        .donut-layout {
            display:flex;
            align-items:center;
            gap:28px;
            min-height:215px;
        }

        .donut {
            width:180px;
            height:180px;
            border-radius:50%;
            display:flex;
            align-items:center;
            justify-content:center;
            flex-shrink:0;
        }

        .donut-hole {
            width:110px;
            height:110px;
            border-radius:50%;
            background:#111827;
            display:flex;
            flex-direction:column;
            align-items:center;
            justify-content:center;
        }

        .donut-hole strong {
            font-size:27px;
            color:#f8fafc;
        }

        .donut-hole span {
            font-size:10px;
            color:#718198;
        }

        .legend-list {
            flex:1;
        }

        .legend-row {
            display:grid;
            grid-template-columns:12px 1fr auto;
            gap:8px;
            align-items:center;
            padding:7px 0;
            color:#cbd5e1;
            font-size:10px;
            border-bottom:1px solid rgba(148,163,184,.08);
        }

        .legend-row b {
            color:#f8fafc;
        }

        .legend-dot {
            width:8px;
            height:8px;
            border-radius:50%;
        }

        .chart-empty {
            padding:30px 10px;
            text-align:center;
            color:#718198;
            font-size:12px;
        }


        @media(max-width:900px) {

            .bar-row {
                grid-template-columns:110px 1fr 30px;
            }

            .donut-layout {
                gap:15px;
            }

            .donut {
                width:145px;
                height:145px;
            }

            .donut-hole {
                width:90px;
                height:90px;
            }
        }

    </style>
    """,
    unsafe_allow_html=True,
)


# ============================================================
# FUNÇÕES GERAIS
# ============================================================

def render_html(text):
    """
    Renderiza HTML usando st.html quando disponível.
    """
    if hasattr(st, "html"):
        st.html(text)
    else:
        st.markdown(text, unsafe_allow_html=True)


def esc(value):
    """
    Escapa valores para HTML.
    """
    return html.escape(
        str(value if value is not None else "")
    )


def formatar_data_br(data_str):
    """
    Converte YYYY-MM-DD para DD/MM/YYYY.
    """
    if not data_str:
        return ""

    texto = str(data_str)

    if texto.lower() in ("nan", "none", ""):
        return ""

    try:
        return datetime.datetime.strptime(
            texto.split("T")[0],
            "%Y-%m-%d",
        ).strftime("%d/%m/%Y")

    except Exception:
        return texto


def converter_para_data(data_str):
    """
    Converte valor para datetime.date.
    """
    if not data_str:
        return None

    try:
        return datetime.date.fromisoformat(
            str(data_str).split("T")[0]
        )

    except Exception:
        return None


# ============================================================
# STATUS DO CARTÃO
# ============================================================

def status_inicial_pelas_datas(carga):
    """
    Define a posição inicial do cartão.

    IMPORTANTE:
    Não grava nada no Firestore.

    Se a carga já possui um status antigo salvo no banco,
    ele é usado somente como posição inicial.

    Depois que o usuário movimentar o cartão nesta sessão,
    a posição passa a ser controlada pelo session_state.
    """

    status_antigo = carga.get("status")

    if status_antigo in STATUS:
        return status_antigo

    hoje = datetime.date.today()

    data_carga = converter_para_data(
        carga.get("data_carga")
    )

    data_saida = converter_para_data(
        carga.get("data_saida")
    )

    data_entrega = converter_para_data(
        carga.get("data_entrega")
    )

    if data_entrega and hoje >= data_entrega:
        return "Entregue / Concluído"

    if data_saida and hoje >= data_saida:
        return "Em Trânsito / Viagem Iniciada"

    if data_carga and hoje >= data_carga:
        return "Carregado / No Pátio"

    return "Aguardando Carregamento"


def inicializar_status_sessao(cargas):
    """
    Inicializa os status dos cartões apenas na sessão atual.

    Nenhum status é salvo no Firestore.
    """

    if "status_cartoes" not in st.session_state:
        st.session_state["status_cartoes"] = {}

    for carga in cargas:
        carga_id = str(carga.get("id", "")).strip()

        if not carga_id:
            continue

        if carga_id not in st.session_state["status_cartoes"]:
            st.session_state["status_cartoes"][carga_id] = (
                status_inicial_pelas_datas(carga)
            )


def obter_status_cartao(carga):
    """
    Retorna o status atual do cartão.
    """

    carga_id = str(carga.get("id", "")).strip()

    if not carga_id:
        return STATUS[0]

    inicializar_status_sessao([carga])

    status = st.session_state["status_cartoes"].get(
        carga_id
    )

    if status not in STATUS:
        status = status_inicial_pelas_datas(carga)
        st.session_state["status_cartoes"][carga_id] = status

    return status


def alterar_status_cartao(carga_id, novo_status):
    """
    Move o cartão somente na sessão.

    NÃO salva no Firebase.
    """

    carga_id = str(carga_id)

    if novo_status not in STATUS:
        return

    st.session_state["status_cartoes"][carga_id] = novo_status


def avancar_status(carga_id):
    """
    Avança o cartão uma coluna.
    """

    status_atual = st.session_state["status_cartoes"].get(
        str(carga_id),
        STATUS[0],
    )

    try:
        indice = STATUS.index(status_atual)
    except ValueError:
        indice = 0

    if indice < len(STATUS) - 1:
        st.session_state["status_cartoes"][str(carga_id)] = (
            STATUS[indice + 1]
        )


def voltar_status(carga_id):
    """
    Volta o cartão uma coluna.
    """

    status_atual = st.session_state["status_cartoes"].get(
        str(carga_id),
        STATUS[0],
    )

    try:
        indice = STATUS.index(status_atual)
    except ValueError:
        indice = 0

    if indice > 0:
        st.session_state["status_cartoes"][str(carga_id)] = (
            STATUS[indice - 1]
        )


# ============================================================
# FIRESTORE
# ============================================================

def firestore_value(v):

    if "stringValue" in v:
        return v["stringValue"]

    if "integerValue" in v:
        return int(v["integerValue"])

    if "doubleValue" in v:
        return float(v["doubleValue"])

    if "booleanValue" in v:
        return bool(v["booleanValue"])

    if "nullValue" in v:
        return None

    if "timestampValue" in v:
        return v["timestampValue"]

    if "referenceValue" in v:
        return v["referenceValue"]

    if "arrayValue" in v:
        return [
            firestore_value(x)
            for x in v["arrayValue"].get("values", [])
        ]

    if "mapValue" in v:
        return {
            k: firestore_value(x)
            for k, x in v["mapValue"].get(
                "fields",
                {}
            ).items()
        }

    return None


def carregar_colecao(colecao):

    documentos = []
    page_token = None

    try:

        while True:

            params = {
                "pageSize": 1000
            }

            if page_token:
                params["pageToken"] = page_token

            response = requests.get(
                f"{FIREBASE_BASE}/{colecao}",
                params=params,
                timeout=15,
            )

            if response.status_code != 200:

                st.error(
                    f"Não foi possível carregar '{colecao}'. "
                    f"Firebase HTTP {response.status_code}."
                )

                return documentos

            data = response.json()

            for doc in data.get(
                "documents",
                []
            ):

                doc_id = doc["name"].split("/")[-1]

                item = {
                    "id": doc_id
                }

                for key, value in doc.get(
                    "fields",
                    {}
                ).items():

                    item[key] = firestore_value(value)

                documentos.append(item)

            page_token = data.get(
                "nextPageToken"
            )

            if not page_token:
                break

        return documentos

    except requests.exceptions.Timeout:

        st.error(
            f"Tempo esgotado ao consultar '{colecao}'."
        )

    except requests.exceptions.RequestException as e:

        st.error(
            f"Erro de conexão com o Firebase: {e}"
        )

    except Exception as e:

        st.error(
            f"Erro ao carregar '{colecao}': {e}"
        )

    return documentos


def firestore_encode(v):

    if isinstance(v, bool):
        return {
            "booleanValue": v
        }

    if isinstance(v, int) and not isinstance(v, bool):
        return {
            "integerValue": str(v)
        }

    if isinstance(v, float):
        return {
            "doubleValue": v
        }

    if v is None:
        return {
            "nullValue": None
        }

    if isinstance(v, list):

        return {
            "arrayValue": {
                "values": [
                    firestore_encode(x)
                    for x in v
                ]
            }
        }

    if isinstance(v, dict):

        return {
            "mapValue": {
                "fields": {
                    k: firestore_encode(x)
                    for k, x in v.items()
                }
            }
        }

    return {
        "stringValue": str(v)
    }


def salvar_documento(
    colecao,
    doc_id,
    dados,
):
    """
    Salva os dados da carga.

    O campo status é deliberadamente ignorado.
    """

    try:

        fields = {}

        for k, v in dados.items():

            if k in ("id", "status"):
                continue

            fields[k] = firestore_encode(v)

        response = requests.patch(
            f"{FIREBASE_BASE}/{colecao}/{doc_id}",
            json={
                "fields": fields
            },
            timeout=15,
        )

        if response.status_code not in (
            200,
            201,
        ):

            st.error(
                "Erro ao salvar no Firebase. "
                f"HTTP {response.status_code}: "
                f"{response.text[:300]}"
            )

            return False

        return True

    except requests.exceptions.Timeout:

        st.error(
            "Tempo esgotado ao salvar no Firebase."
        )

    except requests.exceptions.RequestException as e:

        st.error(
            f"Erro de conexão com o Firebase: {e}"
        )

    except Exception as e:

        st.error(
            f"Erro ao salvar no Firebase: {e}"
        )

    return False


def deletar_documento(
    colecao,
    doc_id,
):

    try:

        response = requests.delete(
            f"{FIREBASE_BASE}/{colecao}/{doc_id}",
            timeout=15,
        )

        if response.status_code not in (
            200,
            204,
        ):

            st.error(
                "Erro ao excluir documento. "
                f"HTTP {response.status_code}."
            )

            return False

        return True

    except requests.exceptions.Timeout:

        st.error(
            "Tempo esgotado ao excluir do Firebase."
        )

    except requests.exceptions.RequestException as e:

        st.error(
            f"Erro de conexão com o Firebase: {e}"
        )

    except Exception as e:

        st.error(
            f"Erro ao excluir do Firebase: {e}"
        )

    return False


def atualizar_dados():

    st.session_state["cargas"] = carregar_colecao(
        "cargas"
    )

    st.session_state["motoristas"] = carregar_colecao(
        "motoristas"
    )

    st.session_state["ajudantes"] = carregar_colecao(
        "ajudantes"
    )

    # Remove status de cargas que não existem mais.
    ids_validos = {
        str(c.get("id"))
        for c in st.session_state["cargas"]
    }

    if "status_cartoes" in st.session_state:

        st.session_state["status_cartoes"] = {
            k: v
            for k, v in st.session_state[
                "status_cartoes"
            ].items()
            if k in ids_validos
        }

    inicializar_status_sessao(
        st.session_state["cargas"]
    )

    st.session_state[
        "ultima_atualizacao"
    ] = datetime.datetime.now().strftime(
        "%d/%m/%Y %H:%M:%S"
    )


# ============================================================
# REGRAS OPERACIONAIS
# ============================================================

def carga_atrasada(carga):

    status = obter_status_cartao(carga)

    if status == "Entregue / Concluído":
        return False

    data = converter_para_data(
        carga.get("data_entrega")
    )

    return bool(
        data and data < datetime.date.today()
    )


def carga_saida_hoje(carga):

    data = converter_para_data(
        carga.get("data_saida")
    )

    return bool(
        data and data == datetime.date.today()
    )


def carga_entrega_hoje(carga):

    data = converter_para_data(
        carga.get("data_entrega")
    )

    return bool(
        data and data == datetime.date.today()
    )


def dias_para_entrega(carga):

    data = converter_para_data(
        carga.get("data_entrega")
    )

    if not data:
        return None

    return (
        data -
        datetime.date.today()
    ).days


def texto_prazo(carga):

    dias = dias_para_entrega(carga)

    if dias is None:
        return ""

    if carga_atrasada(carga):

        return (
            f"🔴 {abs(dias)} dia(s) atrasada"
        )

    if dias == 0:
        return "🟠 Entrega hoje"

    if dias == 1:
        return "🟡 Entrega amanhã"

    return f"🟢 {dias} dias para entrega"


# ============================================================
# DATAFRAME
# ============================================================

def preparar_dataframe(cargas):

    registros = []

    for carga in cargas:

        status = obter_status_cartao(carga)

        ajudantes = carga.get(
            "ajudantes",
            []
        )

        if isinstance(
            ajudantes,
            list,
        ):

            ajudantes_txt = ", ".join(
                map(str, ajudantes)
            )

        else:

            ajudantes_txt = str(
                ajudantes or ""
            )

        registros.append(
            {
                "id": carga.get(
                    "id",
                    ""
                ),

                "motorista": carga.get(
                    "motorista",
                    ""
                ),

                "destino": carga.get(
                    "destino",
                    ""
                ),

                "observacoes": carga.get(
                    "observacoes",
                    ""
                ),

                "ajudantes": ajudantes_txt,

                "data_carga": formatar_data_br(
                    carga.get(
                        "data_carga"
                    )
                ),

                "data_saida": formatar_data_br(
                    carga.get(
                        "data_saida"
                    )
                ),

                "data_entrega": formatar_data_br(
                    carga.get(
                        "data_entrega"
                    )
                ),

                "status": status,
            }
        )

    return pd.DataFrame(
        registros,
        columns=[
            "id",
            "motorista",
            "destino",
            "observacoes",
            "ajudantes",
            "data_carga",
            "data_saida",
            "data_entrega",
            "status",
        ],
    )


# ============================================================
# EXCEL
# ============================================================

def gerar_excel_profissional(df):

    output = io.BytesIO()

    wb = openpyxl.Workbook()

    ws = wb.active

    ws.title = "Relatório de Cargas"

    header_font = Font(
        name="Arial",
        size=10,
        bold=True,
        color="FFFFFF",
    )

    header_fill = PatternFill(
        start_color="2F75B5",
        end_color="2F75B5",
        fill_type="solid",
    )

    data_font = Font(
        name="Arial",
        size=9,
    )

    border = Border(
        left=Side(
            style="thin",
            color="D9D9D9",
        ),
        right=Side(
            style="thin",
            color="D9D9D9",
        ),
        top=Side(
            style="thin",
            color="D9D9D9",
        ),
        bottom=Side(
            style="thin",
            color="D9D9D9",
        ),
    )

    headers = [
        "ID Planejamento",
        "Motorista",
        "Destino",
        "Observações",
        "Ajudantes",
        "Data Carga",
        "Data Saída",
        "Data Entrega",
        "Status",
    ]

    for col_num, header in enumerate(
        headers,
        1,
    ):

        cell = ws.cell(
            1,
            col_num,
            header,
        )

        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )
        cell.border = border

    for row_num, row in enumerate(
        df.values,
        2,
    ):

        for col_num, value in enumerate(
            row,
            1,
        ):

            cell = ws.cell(
                row_num,
                col_num,
                "" if pd.isna(value)
                else value,
            )

            cell.font = data_font
            cell.border = border

            cell.alignment = Alignment(
                horizontal=(
                    "center"
                    if col_num in (
                        1,
                        6,
                        7,
                        8,
                    )
                    else "left"
                ),
                vertical="center",
            )

    for col in ws.columns:

        max_len = max(
            len(
                str(
                    c.value
                    or ""
                )
            )
            for c in col
        )

        letter = get_column_letter(
            col[0].column
        )

        ws.column_dimensions[
            letter
        ].width = min(
            max(
                max_len + 3,
                14,
            ),
            45,
        )

    ws.freeze_panes = "A2"

    ws.auto_filter.ref = ws.dimensions

    wb.save(output)

    return output.getvalue()


# ============================================================
# PDF
# ============================================================

def pdf_text(value):

    texto = str(
        value if value is not None
        else ""
    )

    substituicoes = {
        "🚚": "",
        "📦": "",
        "⏳": "",
        "✅": "",
        "🔴": "",
        "🟠": "",
        "🟡": "",
        "🟢": "",
        "📌": "",
        "👥": "",
        "📝": "",
    }

    for antigo, novo in substituicoes.items():
        texto = texto.replace(
            antigo,
            novo,
        )

    return texto.encode(
        "latin-1",
        errors="replace",
    ).decode(
        "latin-1"
    )


def gerar_pdf(df):

    pdf = FPDF(
        orientation="L",
        unit="mm",
        format="A4",
    )

    pdf.add_page()

    pdf.set_font(
        "Arial",
        "B",
        14,
    )

    pdf.cell(
        277,
        8,
        "Relatorio de Cargas",
        ln=True,
        align="C",
    )

    pdf.set_font(
        "Arial",
        "",
        9,
    )

    pdf.cell(
        277,
        5,
        pdf_text(
            f"Data de geracao: "
            f"{datetime.date.today():%d/%m/%Y}"
        ),
        ln=True,
        align="C",
    )

    pdf.ln(4)

    headers = [
        "ID",
        "Motorista",
        "Destino",
        "Ajudantes",
        "Carga",
        "Saida",
        "Entrega",
        "Status",
    ]

    widths = [
        25,
        40,
        50,
        45,
        28,
        28,
        28,
        33,
    ]

    pdf.set_font(
        "Arial",
        "B",
        8,
    )

    pdf.set_fill_color(
        47,
        117,
        181,
    )

    pdf.set_text_color(
        255,
        255,
        255,
    )

    for width, header in zip(
        widths,
        headers,
    ):

        pdf.cell(
            width,
            7,
            pdf_text(header),
            1,
            0,
            "C",
            True,
        )

    pdf.ln()

    pdf.set_font(
        "Arial",
        "",
        7,
    )

    pdf.set_text_color(
        0,
        0,
        0,
    )

    for _, row in df.iterrows():

        vals = [
            str(
                row.get(
                    "id",
                    "",
                )
            ),

            str(
                row.get(
                    "motorista",
                    "",
                )
            )[:22],

            str(
                row.get(
                    "destino",
                    "",
                )
            )[:30],

            str(
                row.get(
                    "ajudantes",
                    "",
                )
            )[:25],

            str(
                row.get(
                    "data_carga",
                    "",
                )
            ),

            str(
                row.get(
                    "data_saida",
                    "",
                )
            ),

            str(
                row.get(
                    "data_entrega",
                    "",
                )
            ),

            str(
                row.get(
                    "status",
                    "",
                )
            )[:24],
        ]

        for width, value in zip(
            widths,
            vals,
        ):

            pdf.cell(
                width,
                6,
                pdf_text(value),
                1,
                0,
                (
                    "C"
                    if width < 30
                    else "L"
                ),
            )

        pdf.ln()

    return bytes(
        pdf.output(
            dest="S"
        )
    )


# ============================================================
# GRÁFICOS
# ============================================================

def grafico_status_html(
    contagem,
    total,
):

    if not total:

        return (
            '<div class="chart-card">'
            '<div class="chart-empty">'
            "Nenhuma carga encontrada."
            "</div>"
            "</div>"
        )

    partes = []

    inicio = 0

    for status in STATUS:

        valor = contagem.get(
            status,
            0,
        )

        if valor:

            fim = (
                inicio
                + valor / total * 100
            )

            partes.append(
                f"{STATUS_COLORS[status]} "
                f"{inicio:.2f}% "
                f"{fim:.2f}%"
            )

            inicio = fim

    legenda = "".join(
        f"""
        <div class="legend-row">
            <span
                class="legend-dot"
                style="
                    background:
                    {STATUS_COLORS[s]}
                ">
            </span>

            <span>{esc(s)}</span>

            <b>
                {contagem.get(s, 0)}
            </b>
        </div>
        """
        for s in STATUS
        if contagem.get(s, 0)
    )

    return f"""
    <div class="chart-card">

        <div class="chart-heading">
            Cargas por Status

            <small>
                Distribuição atual
            </small>
        </div>

        <div class="donut-layout">

            <div
                class="donut"
                style="
                    background:
                    conic-gradient(
                        {','.join(partes)}
                    )
                "
            >

                <div class="donut-hole">

                    <strong>
                        {total}
                    </strong>

                    <span>
                        cargas
                    </span>

                </div>

            </div>

            <div class="legend-list">
                {legenda}
            </div>

        </div>

    </div>
    """


def grafico_barras_html(
    titulo,
    dados,
    limite=8,
):

    pares = sorted(
        dados.items(),
        key=lambda x: x[1],
        reverse=True,
    )[:limite]

    if not pares:

        return f"""
        <div class="chart-card">

            <div class="chart-heading">
                {esc(titulo)}
            </div>

            <div class="chart-empty">
                Nenhum dado disponível.
            </div>

        </div>
        """

    maximo = max(
        v for _, v in pares
    ) or 1

    linhas = ""

    for nome, valor in pares:

        largura = max(
            7,
            valor / maximo * 100,
        )

        linhas += f"""
        <div class="bar-row">

            <div
                class="bar-label"
                title="{esc(nome)}"
            >
                {esc(nome)}
            </div>

            <div class="bar-track">

                <div
                    class="bar-fill"
                    style="
                        width:
                        {largura:.1f}%
                    "
                >
                </div>

            </div>

            <div class="bar-value">
                {valor}
            </div>

        </div>
        """

    return f"""
    <div class="chart-card">

        <div class="chart-heading">

            {esc(titulo)}

            <small>
                Top {len(pares)}
            </small>

        </div>

        <div class="bar-list">
            {linhas}
        </div>

    </div>
    """


# ============================================================
# CARGA INICIAL
# ============================================================

if "cargas" not in st.session_state:

    st.session_state["cargas"] = (
        carregar_colecao("cargas")
    )


if "motoristas" not in st.session_state:

    st.session_state["motoristas"] = (
        carregar_colecao("motoristas")
    )


if "ajudantes" not in st.session_state:

    st.session_state["ajudantes"] = (
        carregar_colecao("ajudantes")
    )


if "status_cartoes" not in st.session_state:

    st.session_state["status_cartoes"] = {}


if "ultima_atualizacao" not in st.session_state:

    st.session_state[
        "ultima_atualizacao"
    ] = datetime.datetime.now().strftime(
        "%d/%m/%Y %H:%M:%S"
    )


cargas_lista = st.session_state[
    "cargas"
]


# Inicializa status somente na sessão.
inicializar_status_sessao(
    cargas_lista
)


motoristas_lista = [
    m.get(
        "nome",
        "",
    )

    for m in st.session_state[
        "motoristas"
    ]

    if m.get("nome")
]


ajudantes_lista = [
    a.get(
        "nome",
        "",
    )

    for a in st.session_state[
        "ajudantes"
    ]

    if a.get("nome")
]


# ============================================================
# CABEÇALHO
# ============================================================

st.markdown(
    """
    <div class="app-kicker">
        LOGÍSTICA INTELIGENTE
    </div>

    <div class="app-title">
        🚚 Gestão de Cargas
    </div>

    <div class="app-subtitle">
        Torre de controle operacional e
        acompanhamento de planejamentos
    </div>
    """,
    unsafe_allow_html=True,
)


menu = st.radio(
    "Menu Principal",
    [
        "📋 Painel (Kanban)",
        "➕ Nova Carga",
        "👥 Cadastros (Equipe)",
        "📈 Relatórios",
    ],
    horizontal=True,
)


col_up1, col_up2 = st.columns(
    [6, 1]
)


with col_up1:

    st.caption(
        "🕐 Última atualização: "
        f"{st.session_state['ultima_atualizacao']}"
    )


with col_up2:

    if st.button(
        "🔄 Atualizar",
        use_container_width=True,
    ):

        atualizar_dados()

        st.rerun()


st.markdown("---")


# ============================================================
# PAINEL KANBAN
# ============================================================

if menu == "📋 Painel (Kanban)":

    st.subheader(
        "📊 Torre de Controle da Operação"
    )

    total = len(
        cargas_lista
    )


    # --------------------------------------------------------
    # CONTAGEM DOS STATUS DA SESSÃO
    # --------------------------------------------------------

    contagem = {
        status: 0
        for status in STATUS
    }

    for carga in cargas_lista:

        status = obter_status_cartao(
            carga
        )

        if status in contagem:
            contagem[status] += 1


    atrasadas = sum(
        carga_atrasada(c)
        for c in cargas_lista
    )

    saidas_hoje = sum(
        carga_saida_hoje(c)
        for c in cargas_lista
    )

    entregas_hoje = sum(
        carga_entrega_hoje(c)
        for c in cargas_lista
    )


    # --------------------------------------------------------
    # MÉTRICAS
    # --------------------------------------------------------

    m1, m2, m3, m4, m5 = st.columns(
        5
    )

    metricas = [
        (
            "TOTAL DE CARGAS",
            total,
            "Planejamentos cadastrados",
            "blue",
        ),

        (
            "AGUARDANDO",
            contagem[
                STATUS[0]
            ],
            "Aguardando carregamento",
            "yellow",
        ),

        (
            "EM TRÂNSITO",
            contagem[
                STATUS[2]
            ],
            "Viagens iniciadas",
            "green",
        ),

        (
            "ENTREGUES",
            contagem[
                STATUS[3]
            ],
            "Operações concluídas",
            "purple",
        ),

        (
            "ATRASADAS",
            atrasadas,
            "Precisam de atenção",
            "red",
        ),
    ]


    for col, (
        titulo,
        valor,
        sub,
        cor,
    ) in zip(
        (
            m1,
            m2,
            m3,
            m4,
            m5,
        ),
        metricas,
    ):

        with col:

            st.markdown(
                f"""
                <div
                    class="metric-card
                           metric-{cor}"
                >

                    <div
                        class="metric-title"
                    >
                        {titulo}
                    </div>

                    <div
                        class="metric-value"
                    >
                        {valor}
                    </div>

                    <div
                        class="metric-subtitle"
                    >
                        {sub}
                    </div>

                </div>
                """,
                unsafe_allow_html=True,
            )


    # --------------------------------------------------------
    # ALERTAS
    # --------------------------------------------------------

    if atrasadas:

        st.markdown(
            f"""
            <div
                class="alert-box alert-red"
            >
                🔴 <b>Atenção:</b>
                {atrasadas} carga(s)
                estão com prazo
                de entrega vencido.
            </div>
            """,
            unsafe_allow_html=True,
        )


    if saidas_hoje:

        st.markdown(
            f"""
            <div
                class="alert-box alert-yellow"
            >
                🚚 <b>Saídas de hoje:</b>
                {saidas_hoje} carga(s).
            </div>
            """,
            unsafe_allow_html=True,
        )


    if entregas_hoje:

        st.markdown(
            f"""
            <div
                class="alert-box alert-yellow"
            >
                📦 <b>Entregas de hoje:</b>
                {entregas_hoje} carga(s).
            </div>
            """,
            unsafe_allow_html=True,
        )


    if (
        not atrasadas
        and not saidas_hoje
        and not entregas_hoje
        and total
    ):

        st.markdown(
            """
            <div
                class="alert-box alert-green"
            >
                🟢 <b>Operação normal:</b>
                nenhuma carga atrasada
                ou com alerta para hoje.
            </div>
            """,
            unsafe_allow_html=True,
        )


    # --------------------------------------------------------
    # FILTROS
    # --------------------------------------------------------

    st.markdown(
        "### 🔎 Filtros da Operação"
    )

    f1, f2, f3, f4 = st.columns(
        [2, 1.5, 1.5, 2]
    )


    with f1:

        motorista_filtro = st.selectbox(
            "Motorista",
            [
                "Todos os Motoristas"
            ]
            + motoristas_lista,
            key="kanban_motorista",
        )


    with f2:

        data_ini = st.date_input(
            "Data inicial",
            value=(
                datetime.date.today()
                - datetime.timedelta(
                    days=7
                )
            ),
            key="kanban_data_ini",
        )


    with f3:

        data_fim = st.date_input(
            "Data final",
            value=(
                datetime.date.today()
                + datetime.timedelta(
                    days=30
                )
            ),
            key="kanban_data_fim",
        )


    with f4:

        pesquisa = st.text_input(
            "Pesquisar",
            placeholder=(
                "ID, motorista, "
                "destino ou observação..."
            ),
            key="kanban_pesquisa",
        )


    # --------------------------------------------------------
    # FILTRO DAS CARGAS
    # --------------------------------------------------------

    cargas_filtradas = []


    if data_ini > data_fim:

        st.error(
            "A data inicial não pode "
            "ser maior que a data final."
        )

    else:

        termo = (
            pesquisa
            .lower()
            .strip()
        )


        for carga in cargas_lista:

            data_ref = converter_para_data(
                carga.get(
                    "data_saida"
                )
                or carga.get(
                    "data_carga"
                )
            )


            dentro_periodo = (
                data_ref is None
                or data_ini
                <= data_ref
                <= data_fim
            )


            motorista_ok = (
                motorista_filtro
                == "Todos os Motoristas"
                or carga.get(
                    "motorista"
                )
                == motorista_filtro
            )


            texto_busca = " ".join(
                str(
                    carga.get(
                        campo,
                        "",
                    )
                )

                for campo in (
                    "id",
                    "motorista",
                    "destino",
                    "observacoes",
                )
            ).lower()


            busca_ok = (
                not termo
                or termo
                in texto_busca
            )


            if (
                dentro_periodo
                and motorista_ok
                and busca_ok
            ):

                cargas_filtradas.append(
                    carga
                )


    st.caption(
        f"🔎 {len(cargas_filtradas)} "
        "carga(s) encontrada(s)."
    )


    # --------------------------------------------------------
    # KANBAN
    # --------------------------------------------------------

    kanban_cols = st.columns(
        4,
        gap="small",
    )


    for idx, status_coluna in enumerate(
        STATUS
    ):

        with kanban_cols[idx]:

            lista_status = [
                carga
                for carga in cargas_filtradas
                if obter_status_cartao(
                    carga
                )
                == status_coluna
            ]


            # Cabeçalho da coluna
            st.markdown(
                f"""
                <div
                    class="kanban-header"
                    style="
                        border-top:
                        3px solid
                        {STATUS_COLORS[
                            status_coluna
                        ]};
                    "
                >

                    {STATUS_ICONS[
                        status_coluna
                    ]}
                    {status_coluna}

                    <div
                        class="kanban-count"
                    >
                        {len(lista_status)}
                        carga(s)
                    </div>

                </div>
                """,
                unsafe_allow_html=True,
            )


            if not lista_status:

                st.markdown(
                    """
                    <div
                        class="kanban-empty"
                    >
                        Nenhuma carga
                    </div>
                    """,
                    unsafe_allow_html=True,
                )


            # ------------------------------------------------
            # CARTÕES
            # ------------------------------------------------

            for carga in lista_status:

                carga_id = str(
                    carga.get(
                        "id",
                        "",
                    )
                )

                motorista = str(
                    carga.get(
                        "motorista",
                        "",
                    )
                )

                destino = str(
                    carga.get(
                        "destino",
                        "",
                    )
                )

                observacoes = str(
                    carga.get(
                        "observacoes",
                        "",
                    )
                ).strip()

                ajudantes = carga.get(
                    "ajudantes",
                    [],
                )


                if not isinstance(
                    ajudantes,
                    list,
                ):

                    ajudantes = (
                        [ajudantes]
                        if ajudantes
                        else []
                    )


                ajudantes_txt = ", ".join(
                    map(
                        str,
                        ajudantes,
                    )
                )


                atrasada = (
                    carga_atrasada(
                        carga
                    )
                )

                saida_hoje = (
                    carga_saida_hoje(
                        carga
                    )
                )

                entrega_hoje = (
                    carga_entrega_hoje(
                        carga
                    )
                )


                # ------------------------------------------------
                # BADGES
                # ------------------------------------------------

                if atrasada:

                    badge = (
                        '<span '
                        'class="badge badge-red">'
                        '🔴 ATRASADA'
                        '</span>'
                    )

                    borda = "#ef4444"

                elif entrega_hoje:

                    badge = (
                        '<span '
                        'class="badge badge-yellow">'
                        '📦 ENTREGA HOJE'
                        '</span>'
                    )

                    borda = "#f59e0b"

                elif saida_hoje:

                    badge = (
                        '<span '
                        'class="badge badge-yellow">'
                        '🚚 SAI HOJE'
                        '</span>'
                    )

                    borda = "#f59e0b"

                else:

                    badge = (
                        '<span '
                        'class="badge badge-green">'
                        '✓ NO PRAZO'
                        '</span>'
                    )

                    borda = STATUS_COLORS.get(
                        status_coluna,
                        "#64748b",
                    )


                prazo = texto_prazo(
                    carga
                )


                if prazo:

                    badge += (
                        '<span '
                        'class="badge badge-blue">'
                        f'{esc(prazo)}'
                        '</span>'
                    )


                # ------------------------------------------------
                # INFORMAÇÕES EXTRAS
                # ------------------------------------------------

                extra = ""


                if ajudantes_txt:

                    extra += (
                        '<div '
                        'class="card-meta">'
                        '👥 Ajudantes: '
                        f'<strong>'
                        f'{esc(ajudantes_txt)}'
                        '</strong>'
                        '</div>'
                    )


                if observacoes:

                    extra += (
                        '<div '
                        'class="card-meta">'
                        '📝 Obs.: '
                        f'<strong>'
                        f'{esc(observacoes)}'
                        '</strong>'
                        '</div>'
                    )


                # ------------------------------------------------
                # CARTÃO
                # ------------------------------------------------

                render_html(
                    f"""
                    <div
                        class="kanban-card"
                        style="
                            border-left:
                            3px solid
                            {borda};
                        "
                    >

                        <div>
                            {badge}
                        </div>

                        <div
                            class="card-id"
                        >
                            📌 PLANEJAMENTO
                            #{esc(carga_id)}
                        </div>

                        <div
                            class="card-driver"
                        >
                            🚚 {esc(motorista)}
                        </div>

                        <div
                            class="card-label"
                        >
                            Destino
                        </div>

                        <div
                            class=
                            "card-destination"
                        >
                            {esc(destino)}
                        </div>

                        <div
                            class="card-divider"
                        >
                        </div>

                        <div
                            class="card-meta"
                        >
                            📅 Carga:
                            <strong>
                                {esc(
                                    formatar_data_br(
                                        carga.get(
                                            "data_carga"
                                        )
                                    )
                                    or "—"
                                )}
                            </strong>
                        </div>

                        <div
                            class="card-meta"
                        >
                            🚚 Saída:
                            <strong>
                                {esc(
                                    formatar_data_br(
                                        carga.get(
                                            "data_saida"
                                        )
                                    )
                                    or "—"
                                )}
                            </strong>
                        </div>

                        <div
                            class="card-meta"
                        >
                            📦 Entrega:
                            <strong>
                                {esc(
                                    formatar_data_br(
                                        carga.get(
                                            "data_entrega"
                                        )
                                    )
                                    or "—"
                                )}
                            </strong>
                        </div>

                        {extra}

                    </div>
                    """,
                )


                # ------------------------------------------------
                # BOTÕES DE MOVIMENTAÇÃO
                # ------------------------------------------------

                indice_status = STATUS.index(
                    status_coluna
                )


                b1, b2 = st.columns(
                    2,
                    gap="small",
                )


                with b1:

                    if indice_status > 0:

                        if st.button(
                            "⬅️ Voltar",
                            key=(
                                f"voltar_"
                                f"{carga_id}"
                            ),
                            use_container_width=True,
                        ):

                            voltar_status(
                                carga_id
                            )

                            st.rerun()

                    else:

                        st.button(
                            "⬅️ Voltar",
                            key=(
                                f"voltar_disabled_"
                                f"{carga_id}"
                            ),
                            disabled=True,
                            use_container_width=True,
                        )


                with b2:

                    if indice_status < len(
                        STATUS
                    ) - 1:

                        if st.button(
                            "Avançar ➡️",
                            key=(
                                f"avancar_"
                                f"{carga_id}"
                            ),
                            type="primary",
                            use_container_width=True,
                        ):

                            avancar_status(
                                carga_id
                            )

                            st.rerun()

                    else:

                        st.button(
                            "Concluído ✅",
                            key=(
                                f"concluido_"
                                f"{carga_id}"
                            ),
                            disabled=True,
                            use_container_width=True,
                        )


                st.caption(
                    "Status do cartão é controlado "
                    "na sessão e não é salvo no Firebase."
                )


                # ------------------------------------------------
                # EDITAR / EXCLUIR
                # ------------------------------------------------

                e1, e2 = st.columns(
                    2,
                    gap="small",
                )


                with e1:

                    if st.button(
                        "✏️ Editar",
                        key=(
                            f"edit_"
                            f"{carga_id}"
                        ),
                        use_container_width=True,
                    ):

                        st.session_state[
                            f"editando_{carga_id}"
                        ] = True

                        st.rerun()


                with e2:

                    if st.button(
                        "🗑️ Excluir",
                        key=(
                            f"delete_"
                            f"{carga_id}"
                        ),
                        use_container_width=True,
                    ):

                        st.session_state[
                            f"confirmar_{carga_id}"
                        ] = True

                        st.rerun()


                # ------------------------------------------------
                # CONFIRMAÇÃO DE EXCLUSÃO
                # ------------------------------------------------

                if st.session_state.get(
                    f"confirmar_{carga_id}",
                    False,
                ):

                    st.warning(
                        f"Excluir o planejamento "
                        f"#{carga_id}?"
                    )


                    x1, x2 = st.columns(
                        2
                    )


                    with x1:

                        if st.button(
                            "Sim, excluir",
                            key=(
                                f"confirm_yes_"
                                f"{carga_id}"
                            ),
                            type="primary",
                            use_container_width=True,
                        ):

                            if deletar_documento(
                                "cargas",
                                carga_id,
                            ):

                                st.session_state[
                                    "cargas"
                                ] = [
                                    x
                                    for x
                                    in st.session_state[
                                        "cargas"
                                    ]

                                    if str(
                                        x.get(
                                            "id"
                                        )
                                    )
                                    != carga_id
                                ]


                                st.session_state[
                                    "status_cartoes"
                                ].pop(
                                    carga_id,
                                    None,
                                )


                                st.rerun()


                    with x2:

                        if st.button(
                            "Cancelar",
                            key=(
                                f"confirm_no_"
                                f"{carga_id}"
                            ),
                            use_container_width=True,
                        ):

                            st.session_state[
                                f"confirmar_{carga_id}"
                            ] = False

                            st.rerun()


                # ------------------------------------------------
                # EDIÇÃO
                # ------------------------------------------------

                if st.session_state.get(
                    f"editando_{carga_id}",
                    False,
                ):

                    with st.form(
                        f"form_edit_{carga_id}"
                    ):

                        st.markdown(
                            f"**✏️ Editando "
                            f"Planejamento "
                            f"#{esc(carga_id)}**"
                        )


                        mot = carga.get(
                            "motorista",
                            "",
                        )


                        if mot in motoristas_lista:

                            mot_idx = (
                                motoristas_lista.index(
                                    mot
                                )
                            )

                        else:

                            mot_idx = 0


                        novo_motorista = st.selectbox(
                            "Motorista",
                            motoristas_lista
                            or [""],
                            index=mot_idx,
                        )


                        novo_destino = st.text_input(
                            "Destino",
                            value=str(
                                carga.get(
                                    "destino",
                                    "",
                                )
                            ),
                        )


                        novas_obs = st.text_area(
                            "Observações / Rota",
                            value=str(
                                carga.get(
                                    "observacoes",
                                    "",
                                )
                            ),
                        )


                        atuais = carga.get(
                            "ajudantes",
                            [],
                        )


                        if not isinstance(
                            atuais,
                            list,
                        ):

                            atuais = []


                        novos_ajudantes = (
                            st.multiselect(
                                "Ajudantes",
                                ajudantes_lista,
                                default=[
                                    x
                                    for x in atuais
                                    if x
                                    in ajudantes_lista
                                ],
                            )
                        )


                        nova_carga = st.date_input(
                            "Data de Carregamento",
                            value=(
                                converter_para_data(
                                    carga.get(
                                        "data_carga"
                                    )
                                )
                                or datetime.date.today()
                            ),
                        )


                        nova_saida = st.date_input(
                            "Data de Saída",
                            value=(
                                converter_para_data(
                                    carga.get(
                                        "data_saida"
                                    )
                                )
                                or datetime.date.today()
                            ),
                        )


                        nova_entrega = st.date_input(
                            "Data de Entrega",
                            value=(
                                converter_para_data(
                                    carga.get(
                                        "data_entrega"
                                    )
                                )
                                or datetime.date.today()
                            ),
                        )


                        salvar_edicao = (
                            st.form_submit_button(
                                "💾 Salvar Alterações",
                                type="primary",
                                use_container_width=True,
                            )
                        )


                        if salvar_edicao:

                            if nova_entrega < nova_saida:

                                st.error(
                                    "A data de entrega "
                                    "não pode ser anterior "
                                    "à data de saída."
                                )

                            elif nova_saida < nova_carga:

                                st.error(
                                    "A data de saída "
                                    "não pode ser anterior "
                                    "à data de carregamento."
                                )

                            else:

                                dados_atualizados = {
                                    "id": carga_id,

                                    "motorista":
                                        novo_motorista,

                                    "destino":
                                        novo_destino,

                                    "observacoes":
                                        novas_obs,

                                    "ajudantes":
                                        novos_ajudantes,

                                    "data_carga":
                                        str(nova_carga),

                                    "data_saida":
                                        str(nova_saida),

                                    "data_entrega":
                                        str(nova_entrega),
                                }


                                if salvar_documento(
                                    "cargas",
                                    carga_id,
                                    dados_atualizados,
                                ):

                                    for i, item in enumerate(
                                        st.session_state[
                                            "cargas"
                                        ]
                                    ):

                                        if str(
                                            item.get(
                                                "id"
                                            )
                                        ) == carga_id:

                                            st.session_state[
                                                "cargas"
                                            ][i] = (
                                                dados_atualizados
                                            )

                                            break


                                    st.session_state[
                                        f"editando_{carga_id}"
                                    ] = False


                                    st.success(
                                        "Alterações "
                                        "salvas com sucesso."
                                    )

                                    st.rerun()


# ============================================================
# NOVA CARGA
# ============================================================

elif menu == "➕ Nova Carga":

    st.subheader(
        "➕ Cadastrar Novo Agendamento de Carga"
    )


    st.info(
        "📌 Informe o número do planejamento "
        "exatamente como recebido do sistema "
        "de montagem de cargas."
    )


    with st.form(
        "form_nova_carga"
    ):

        id_planejamento = st.text_input(
            "Número do Planejamento / ID da Carga",
            placeholder="Ex.: 1042",
        )


        c1, c2 = st.columns(
            2
        )


        with c1:

            motorista = st.selectbox(
                "Motorista Responsável",
                motoristas_lista
                or ["Nenhum cadastrado"],
            )


            destino = st.text_input(
                "Região / Cidades de Destino",
                placeholder=(
                    "Ex.: Uberaba, Araxá"
                ),
            )


            observacoes = st.text_area(
                "Observações / Rota",
                placeholder=(
                    "Ex.: Entregas em "
                    "lojas diferentes"
                ),
            )


        with c2:

            ajudantes = st.multiselect(
                "Ajudantes da Viagem",
                ajudantes_lista,
            )


            data_carga = st.date_input(
                "Data do Carregamento",
                value=datetime.date.today(),
            )


            data_saida = st.date_input(
                "Data de Saída",
                value=datetime.date.today(),
            )


            data_entrega = st.date_input(
                "Data Prevista de Entrega",
                value=datetime.date.today(),
            )


        st.markdown(
            "### 📌 Posição inicial"
        )

        status_inicial = st.selectbox(
            "Onde o cartão deve começar?",
            STATUS,
        )


        salvar = st.form_submit_button(
            "💾 Salvar e Agendar Carga",
            type="primary",
            use_container_width=True,
        )


        if salvar:

            id_planejamento = (
                str(
                    id_planejamento
                ).strip()
            )


            if (
                not id_planejamento
                or not destino
                or not motorista
            ):

                st.error(
                    "Preencha o Número do "
                    "Planejamento, Motorista "
                    "e Região de Destino."
                )


            elif data_saida < data_carga:

                st.error(
                    "A Data de Saída não pode "
                    "ser anterior à Data do "
                    "Carregamento."
                )


            elif data_entrega < data_saida:

                st.error(
                    "A Data Prevista de Entrega "
                    "não pode ser anterior à "
                    "Data de Saída."
                )


            elif any(
                str(
                    c.get("id")
                )
                == id_planejamento

                for c in cargas_lista
            ):

                st.error(
                    "Já existe uma carga "
                    f"com o ID/Planejamento "
                    f"'{id_planejamento}'."
                )


            else:

                nova_carga = {

                    "id":
                        id_planejamento,

                    "motorista":
                        motorista,

                    "destino":
                        destino,

                    "observacoes":
                        observacoes,

                    "ajudantes":
                        ajudantes,

                    "data_carga":
                        str(data_carga),

                    "data_saida":
                        str(data_saida),

                    "data_entrega":
                        str(data_entrega),
                }


                # IMPORTANTE:
                # status não vai para o Firebase.
                #
                # A posição fica somente na sessão.
                #
                # Assim o cartão pode percorrer
                # as colunas sem salvar status.

                if salvar_documento(
                    "cargas",
                    id_planejamento,
                    nova_carga,
                ):

                    st.session_state[
                        "cargas"
                    ].append(
                        nova_carga
                    )


                    st.session_state[
                        "status_cartoes"
                    ][
                        id_planejamento
                    ] = status_inicial


                    st.success(
                        f"✅ Planejamento "
                        f"#{id_planejamento} "
                        "cadastrado!"
                    )

                    st.rerun()


# ============================================================
# CADASTROS
# ============================================================

elif menu == "👥 Cadastros (Equipe)":

    st.subheader(
        "👥 Gerenciamento de Motoristas "
        "e Ajudantes"
    )


    c1, c2 = st.columns(
        2
    )


    # ========================================================
    # MOTORISTAS
    # ========================================================

    with c1:

        st.markdown(
            "### 🚚 Motoristas"
        )


        with st.form(
            "form_cad_mot",
            clear_on_submit=True,
        ):

            nome = st.text_input(
                "Adicionar novo motorista"
            )


            enviar = st.form_submit_button(
                "Cadastrar Motorista",
                use_container_width=True,
            )


            if enviar:

                nome = nome.strip()


                if not nome:

                    st.error(
                        "Informe o nome "
                        "do motorista."
                    )


                elif any(
                    str(
                        m.get(
                            "nome",
                            "",
                        )
                    ).lower()
                    == nome.lower()

                    for m in st.session_state[
                        "motoristas"
                    ]
                ):

                    st.error(
                        "Este motorista "
                        "já está cadastrado."
                    )


                else:

                    doc_id = (
                        "mot_"
                        f"{int("
                        "datetime.datetime.now()"
                        ".timestamp() * 1000"
                        ")}"
                    )


                    dados = {
                        "id": doc_id,
                        "nome": nome,
                    }


                    if salvar_documento(
                        "motoristas",
                        doc_id,
                        dados,
                    ):

                        st.session_state[
                            "motoristas"
                        ].append(
                            dados
                        )

                        st.success(
                            f"Motorista "
                            f"{nome} adicionado!"
                        )

                        st.rerun()


        st.markdown("---")


        if not st.session_state[
            "motoristas"
        ]:

            st.info(
                "Nenhum motorista cadastrado."
            )


        for obj in st.session_state[
            "motoristas"
        ]:

            nome = obj.get(
                "nome",
                "",
            )

            doc_id = obj.get(
                "id",
                nome,
            )


            a, b = st.columns(
                [4, 2]
            )


            a.write(
                f"🚚 {nome}"
            )


            if b.button(
                "Excluir",
                key=f"del_mot_{doc_id}",
                use_container_width=True,
            ):

                st.session_state[
                    f"conf_mot_{doc_id}"
                ] = True

                st.rerun()


            if st.session_state.get(
                f"conf_mot_{doc_id}",
                False,
            ):

                st.warning(
                    f"Excluir o motorista "
                    f"'{nome}'?"
                )


                y, n = st.columns(
                    2
                )


                with y:

                    if st.button(
                        "Confirmar",
                        key=f"yes_mot_{doc_id}",
                        type="primary",
                        use_container_width=True,
                    ):

                        if deletar_documento(
                            "motoristas",
                            doc_id,
                        ):

                            st.session_state[
                                "motoristas"
                            ] = [
                                x
                                for x
                                in st.session_state[
                                    "motoristas"
                                ]

                                if x.get(
                                    "id"
                                )
                                != doc_id
                            ]

                            st.rerun()


                with n:

                    if st.button(
                        "Cancelar",
                        key=f"no_mot_{doc_id}",
                        use_container_width=True,
                    ):

                        st.session_state[
                            f"conf_mot_{doc_id}"
                        ] = False

                        st.rerun()


    # ========================================================
    # AJUDANTES
    # ========================================================

    with c2:

        st.markdown(
            "### 👥 Ajudantes"
        )


        with st.form(
            "form_cad_aju",
            clear_on_submit=True,
        ):

            nome = st.text_input(
                "Adicionar novo ajudante"
            )


            enviar = st.form_submit_button(
                "Cadastrar Ajudante",
                use_container_width=True,
            )


            if enviar:

                nome = nome.strip()


                if not nome:

                    st.error(
                        "Informe o nome "
                        "do ajudante."
                    )


                elif any(
                    str(
                        a.get(
                            "nome",
                            "",
                        )
                    ).lower()
                    == nome.lower()

                    for a in st.session_state[
                        "ajudantes"
                    ]
                ):

                    st.error(
                        "Este ajudante "
                        "já está cadastrado."
                    )


                else:

                    doc_id = (
                        "aju_"
                        f"{int("
                        "datetime.datetime.now()"
                        ".timestamp() * 1000"
                        ")}"
                    )


                    dados = {
                        "id": doc_id,
                        "nome": nome,
                    }


                    if salvar_documento(
                        "ajudantes",
                        doc_id,
                        dados,
                    ):

                        st.session_state[
                            "ajudantes"
                        ].append(
                            dados
                        )

                        st.success(
                            f"Ajudante "
                            f"{nome} adicionado!"
                        )

                        st.rerun()


        st.markdown("---")


        if not st.session_state[
            "ajudantes"
        ]:

            st.info(
                "Nenhum ajudante cadastrado."
            )


        for obj in st.session_state[
            "ajudantes"
        ]:

            nome = obj.get(
                "nome",
                "",
            )

            doc_id = obj.get(
                "id",
                nome,
            )


            a, b = st.columns(
                [4, 2]
            )


            a.write(
                f"👤 {nome}"
            )


            if b.button(
                "Excluir",
                key=f"del_aju_{doc_id}",
                use_container_width=True,
            ):

                st.session_state[
                    f"conf_aju_{doc_id}"
                ] = True

                st.rerun()


            if st.session_state.get(
                f"conf_aju_{doc_id}",
                False,
            ):

                st.warning(
                    f"Excluir o ajudante "
                    f"'{nome}'?"
                )


                y, n = st.columns(
                    2
                )


                with y:

                    if st.button(
                        "Confirmar",
                        key=f"yes_aju_{doc_id}",
                        type="primary",
                        use_container_width=True,
                    ):

                        if deletar_documento(
                            "ajudantes",
                            doc_id,
                        ):

                            st.session_state[
                                "ajudantes"
                            ] = [
                                x
                                for x
                                in st.session_state[
                                    "ajudantes"
                                ]

                                if x.get(
                                    "id"
                                )
                                != doc_id
                            ]

                            st.rerun()


                with n:

                    if st.button(
                        "Cancelar",
                        key=f"no_aju_{doc_id}",
                        use_container_width=True,
                    ):

                        st.session_state[
                            f"conf_aju_{doc_id}"
                        ] = False

                        st.rerun()


# ============================================================
# RELATÓRIOS
# ============================================================

elif menu == "📈 Relatórios":

    st.subheader(
        "📈 Relatórios e Exportação de Dados"
    )


    if not cargas_lista:

        st.info(
            "Nenhuma carga cadastrada "
            "para gerar relatórios."
        )


    else:

        f1, f2, f3 = st.columns(
            3
        )


        with f1:

            filtro_motorista = (
                st.selectbox(
                    "Motorista",
                    ["Todos"]
                    + motoristas_lista,
                    key="rel_motorista",
                )
            )


        with f2:

            filtro_status = (
                st.selectbox(
                    "Status",
                    ["Todos"]
                    + STATUS,
                    key="rel_status",
                )
            )


        with f3:

            filtro_busca = st.text_input(
                "Pesquisar ID, destino ou observação",
                key="rel_busca",
            )


        cargas_relatorio = list(
            cargas_lista
        )


        if filtro_motorista != "Todos":

            cargas_relatorio = [
                c
                for c in cargas_relatorio
                if c.get(
                    "motorista"
                )
                == filtro_motorista
            ]


        if filtro_status != "Todos":

            cargas_relatorio = [
                c
                for c in cargas_relatorio
                if obter_status_cartao(
                    c
                )
                == filtro_status
            ]


        if filtro_busca.strip():

            termo = (
                filtro_busca
                .lower()
                .strip()
            )


            cargas_relatorio = [
                c
                for c in cargas_relatorio

                if termo
                in " ".join(
                    str(
                        c.get(
                            k,
                            "",
                        )
                    )

                    for k in (
                        "id",
                        "motorista",
                        "destino",
                        "observacoes",
                    )
                ).lower()
            ]


        total_r = len(
            cargas_relatorio
        )


        entregues = sum(
            obter_status_cartao(c)
            == "Entregue / Concluído"

            for c in cargas_relatorio
        )


        transito = sum(
            obter_status_cartao(c)
            == "Em Trânsito / Viagem Iniciada"

            for c in cargas_relatorio
        )


        atrasadas_r = sum(
            carga_atrasada(c)
            for c in cargas_relatorio
        )


        patio = sum(
            obter_status_cartao(c)
            == "Carregado / No Pátio"

            for c in cargas_relatorio
        )


        aguardando = sum(
            obter_status_cartao(c)
            == "Aguardando Carregamento"

            for c in cargas_relatorio
        )


        conclusao = (
            entregues
            / total_r
            * 100
            if total_r
            else 0
        )


        # ----------------------------------------------------
        # MÉTRICAS
        # ----------------------------------------------------

        r1, r2, r3, r4, r5 = st.columns(
            5
        )


        dados_metricas = [

            (
                "CARGAS",
                total_r,
                "Resultado do filtro",
                "blue",
            ),

            (
                "ENTREGUES",
                entregues,
                "Concluídas",
                "green",
            ),

            (
                "PÁTIO",
                patio,
                "Carregadas / no pátio",
                "yellow",
            ),

            (
                "EM TRÂNSITO",
                transito,
                "Viagens em andamento",
                "purple",
            ),

            (
                "CONCLUSÃO",
                f"{conclusao:.1f}%",
                "Taxa de entregas",
                "blue",
            ),
        ]


        for col, (
            titulo,
            valor,
            sub,
            cor,
        ) in zip(
            (
                r1,
                r2,
                r3,
                r4,
                r5,
            ),
            dados_metricas,
        ):

            with col:

                st.markdown(
                    f"""
                    <div
                        class=
                        "metric-card
                         metric-{cor}"
                    >

                        <div
                            class=
                            "metric-title"
                        >
                            {titulo}
                        </div>

                        <div
                            class=
                            "metric-value"
                        >
                            {valor}
                        </div>

                        <div
                            class=
                            "metric-subtitle"
                        >
                            {sub}
                        </div>

                    </div>
                    """,
                    unsafe_allow_html=True,
                )


        if atrasadas_r:

            st.markdown(
                f"""
                <div
                    class=
                    "alert-box alert-red"
                >
                    🔴 O filtro atual possui
                    {atrasadas_r}
                    carga(s) atrasada(s).
                </div>
                """,
                unsafe_allow_html=True,
            )


        # ----------------------------------------------------
        # VISÃO GERENCIAL
        # ----------------------------------------------------

        st.markdown(
            '<div class="section-title">'
            '📊 Visão Gerencial'
            '</div>',
            unsafe_allow_html=True,
        )


        status_contagem = {
            s: 0
            for s in STATUS
        }


        motorista_contagem = {}

        evolucao = {}


        for c in cargas_relatorio:

            status = obter_status_cartao(
                c
            )


            if status in status_contagem:

                status_contagem[
                    status
                ] += 1


            motorista = (
                c.get(
                    "motorista"
                )
                or "Sem motorista"
            )


            motorista_contagem[
                motorista
            ] = (
                motorista_contagem.get(
                    motorista,
                    0,
                )
                + 1
            )


            data = converter_para_data(
                c.get(
                    "data_saida"
                )
                or c.get(
                    "data_carga"
                )
            )


            if data:

                evolucao[data] = (
                    evolucao.get(
                        data,
                        0,
                    )
                    + 1
                )


        st.markdown(
            grafico_status_html(
                status_contagem,
                total_r,
            ),
            unsafe_allow_html=True,
        )


        st.markdown(
            grafico_barras_html(
                "Cargas por Motorista",
                motorista_contagem,
                limite=8,
            ),
            unsafe_allow_html=True,
        )


        if evolucao:

            st.markdown(
                '<div class="chart-card">'
                '<div class="chart-heading">'
                '📅 Movimentação por Data'
                '<small>'
                'Saída / carregamento'
                '</small>'
                '</div>',
                unsafe_allow_html=True,
            )


            df_evo = pd.DataFrame(
                [
                    {
                        "Data":
                            d.strftime(
                                "%d/%m/%Y"
                            ),

                        "Cargas":
                            v,
                    }

                    for d, v
                    in sorted(
                        evolucao.items()
                    )
                ]
            )


            st.dataframe(
                df_evo,
                use_container_width=True,
                hide_index=True,
            )


            st.markdown(
                "</div>",
                unsafe_allow_html=True,
            )


        # ----------------------------------------------------
        # DETALHAMENTO
        # ----------------------------------------------------

        st.markdown(
            "### 📋 Detalhamento das Cargas"
        )


        df = preparar_dataframe(
            cargas_relatorio
        )


        if df.empty:

            st.info(
                "Nenhuma carga encontrada "
                "com os filtros atuais."
            )


        else:

            st.dataframe(
                df,
                use_container_width=True,
                hide_index=True,
            )


            # ------------------------------------------------
            # EXPORTAÇÃO
            # ------------------------------------------------

            st.markdown(
                "### 📥 Exportar Arquivos"
            )


            e1, e2 = st.columns(
                2
            )


            with e1:

                st.download_button(
                    "📥 Baixar Excel (.xlsx)",

                    data=(
                        gerar_excel_profissional(
                            df
                        )
                    ),

                    file_name=(
                        "relatorio_de_cargas.xlsx"
                    ),

                    mime=(
                        "application/vnd.openxmlformats-"
                        "officedocument.spreadsheetml.sheet"
                    ),

                    use_container_width=True,
                )


            with e2:

                st.download_button(
                    "📄 Baixar PDF",

                    data=(
                        gerar_pdf(df)
                    ),

                    file_name=(
                        "relatorio_de_cargas.pdf"
                    ),

                    mime="application/pdf",

                    use_container_width=True,
                )

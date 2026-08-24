import datetime
import io
import json
import os
from zoneinfo import ZoneInfo
import requests
import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
from fpdf import FPDF
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================
# CONFIGURAÇÃO DA PÁGINA
# ============================================================

st.set_page_config(
    page_title="Gestão de Cargas - Logística",
    page_icon="🚚",
    layout="wide",
    initial_sidebar_state="expanded",
)


# ============================================================
# CSS
# ============================================================

st.markdown(
    """
    <style>
        #MainMenu {visibility: hidden;}
        header[data-testid="stHeader"] {
            background: transparent !important;
        }
        [data-testid="stToolbar"] {
            background: transparent !important;
        }


        /* MENU FLUTUANTE: oculto, abre ao passar no botão ☰ */
        section[data-testid="stSidebar"] {
            position: fixed !important;
            left: 0 !important;
            top: 0 !important;
            height: 100vh !important;
            width: 0 !important;
            min-width: 0 !important;
            overflow: visible !important;
            z-index: 999999 !important;
            transition: width .24s ease, min-width .24s ease !important;
            background: transparent !important;
            border: 0 !important;
        }

        section[data-testid="stSidebar"]::before {
            content: "☰";
            position: fixed;
            left: 14px;
            top: 18px;
            width: 44px;
            height: 44px;
            display: flex;
            align-items: center;
            justify-content: center;
            border-radius: 11px;
            background: linear-gradient(145deg,#172131,#101827);
            border: 1px solid #2b3950;
            color: #60a5fa;
            font-size: 22px;
            font-weight: 900;
            box-shadow: 0 10px 28px rgba(0,0,0,.28);
            cursor: pointer;
            transition: all .18s ease;
        }

        section[data-testid="stSidebar"]:hover::before {
            left: 296px;
            background: #18263a;
            border-color: #3b82f6;
            color: #dbeafe;
            content: "‹";
        }

        section[data-testid="stSidebar"]:hover {
            width: 285px !important;
            min-width: 285px !important;
        }

        section[data-testid="stSidebar"] > div:first-child {
            width: 285px !important;
            min-width: 285px !important;
            height: 100vh !important;
            padding-top: 18px !important;
            background: linear-gradient(180deg,#0d1522 0%,#0a101a 100%) !important;
            border-right: 1px solid #263449;
            box-shadow: 18px 0 42px rgba(0,0,0,.30);
            overflow-y: auto !important;
            overflow-x: hidden !important;
            transform: translateX(-102%);
            transition: transform .24s ease !important;
        }

        section[data-testid="stSidebar"]:hover > div:first-child {
            transform: translateX(0);
        }

        section[data-testid="stSidebar"] [data-testid="stSidebarCollapseButton"] {
            display: none !important;
        }

        section[data-testid="stSidebar"] div[role="radiogroup"] label {
            border-radius: 9px !important;
            padding: 8px 10px !important;
            margin-bottom: 3px !important;
            transition: background .15s ease, transform .15s ease !important;
        }

        section[data-testid="stSidebar"] div[role="radiogroup"] label:hover {
            background: rgba(59,130,246,.10) !important;
            transform: translateX(2px);
        }

        [data-testid="stAppViewContainer"] > .main {
            margin-left: 0 !important;
        }

        .block-container {
            max-width: 1680px !important;
            padding-left: 78px !important;
            padding-right: 28px !important;
            padding-top: 4.6rem !important;
        }

        .page-head {
            display:flex;
            align-items:flex-start;
            justify-content:space-between;
            gap:16px;
            margin:0 0 20px 0;
            position:relative;
            z-index:10;
        }
        .page-title {
            color:#f8fafc;
            font-size:27px;
            font-weight:880;
            letter-spacing:-.5px;
            line-height:1.05;
        }
        .page-subtitle {color:#8796aa;font-size:12px;margin-top:7px;}
        .date-chip {
            border:1px solid #2b3950;
            background:#141c29;
            border-radius:9px;
            padding:9px 12px;
            color:#cbd5e1;
            font-size:11px;
            white-space:nowrap;
        }

        .kpi-grid {
            display:grid;
            grid-template-columns:repeat(7,minmax(115px,1fr));
            gap:10px;
            margin:6px 0 18px;
        }
        .kpi-card {
            background:linear-gradient(145deg,#151e2c,#101722);
            border:1px solid #263449;
            border-radius:12px;
            padding:14px 14px 13px;
            min-height:112px;
            box-shadow:0 10px 26px rgba(0,0,0,.13);
        }
        .kpi-icon {
            width:32px;height:32px;border-radius:50%;
            display:flex;align-items:center;justify-content:center;
            font-size:15px;margin-bottom:10px;
            background:rgba(59,130,246,.12);
            border:1px solid rgba(59,130,246,.20);
        }
        .kpi-title {color:#a7b3c4;font-size:9px;font-weight:850;letter-spacing:.35px;text-transform:uppercase;}
        .kpi-value {color:#f8fafc;font-size:24px;font-weight:900;margin-top:5px;line-height:1;}
        .kpi-sub {color:#718198;font-size:9px;margin-top:7px;}

        .attention-panel {
            background:linear-gradient(145deg,rgba(36,20,24,.72),rgba(19,20,28,.92));
            border:1px solid rgba(239,68,68,.20);
            border-radius:13px;
            padding:15px 16px 8px;
            margin-bottom:18px;
        }
        .attention-title {
            color:#fb7185;
            font-size:12px;
            font-weight:900;
            letter-spacing:.3px;
            padding-bottom:10px;
            border-bottom:1px solid rgba(148,163,184,.10);
        }
        .attention-row {
            display:grid;
            grid-template-columns:90px minmax(200px,1.8fr) minmax(120px,1fr) minmax(100px,.9fr) minmax(120px,1fr);
            gap:12px;
            align-items:center;
            padding:10px 6px;
            border-bottom:1px solid rgba(148,163,184,.08);
            font-size:10px;
            color:#cbd5e1;
        }
        .attention-row:last-child {border-bottom:0;}
        .attention-id {font-weight:900;color:#f8fafc;font-size:11px;}
        .attention-danger {color:#fb7185;font-weight:800;}
        .attention-warning {color:#f59e0b;font-weight:800;}
        .attention-empty {padding:18px 6px;color:#86efac;font-size:11px;}

        .kanban-shell {
            border:1px solid #263449;
            border-radius:14px;
            padding:10px;
            background:rgba(12,18,28,.55);
            margin-top:8px;
        }
        .kanban-header {
            text-align:left;
            background:transparent !important;
            color:#f8fafc!important;
            padding:10px 8px !important;
            border-radius:9px !important;
            font-weight:800;
            font-size:12px !important;
            border:0 !important;
            border-bottom:1px solid #263449 !important;
            margin-bottom:8px !important;
            box-shadow:none !important;
        }
        .kanban-count {
            float:right;
            color:#a5b3c5!important;
            font-size:10px!important;
            background:#202a3a;
            border-radius:999px;
            padding:2px 7px;
        }
        .kanban-card {
            background:linear-gradient(145deg,#17202e 0%,#121925 100%) !important;
            border:1px solid #2a3749 !important;
            border-radius:10px !important;
            padding:11px 46px 14px 12px !important;
            margin:0 0 8px 0 !important;
            box-shadow:none !important;
        }
        .card-topline {display:flex;justify-content:space-between;align-items:center;gap:8px;margin-bottom:8px;}
        .card-id {color:#f8fafc!important;font-size:11px!important;font-weight:900!important;}
        .card-destination {color:#f8fafc!important;font-size:11px!important;font-weight:750!important;margin:0 0 8px!important;}
        .card-meta {color:#a5b3c5!important;font-size:9px!important;line-height:1.7!important;}
        .card-deadline {margin-top:8px;color:#cbd5e1;font-size:9px;}
        .card-deadline strong {color:#f8fafc;}
        .badge {padding:3px 6px!important;font-size:8px!important;margin:0!important;}

        /* Ação compacta do cartão: seta encaixada no canto inferior direito */
        div[data-testid="stPopover"] > button {
            border-radius:8px !important;
            border:1px solid #304158 !important;
            background:#151f2d !important;
            min-height:32px !important;
            height:32px !important;
            width:32px !important;
            min-width:32px !important;
            padding:0 !important;
            color:#cbd5e1 !important;
            font-size:20px !important;
            font-weight:700 !important;
            line-height:1 !important;
            box-shadow:0 4px 12px rgba(0,0,0,.18) !important;
        }

        div[data-testid="stPopover"] > button:hover {
            border-color:#3b82f6 !important;
            background:#1a2a40 !important;
            color:#ffffff !important;
            transform:translateX(2px);
        }


        .section-caption {color:#f1f5f9;font-size:15px;font-weight:850;margin:10px 0 9px;}
        .app-footer {
            margin-top:26px;
            padding:14px 0 6px;
            border-top:1px solid rgba(148,163,184,.10);
            text-align:center;
            color:#64748b;
            font-size:9px;
        }

        @media (max-width:1100px) {
            .kpi-grid {grid-template-columns:repeat(4,1fr);}
            .attention-row {grid-template-columns:80px 1fr 120px;}
            .attention-row .hide-small {display:none;}
        }
        @media (max-width:768px) {
            .block-container {padding-left:64px!important;padding-right:12px!important;padding-top:4.4rem!important;}
            .kpi-grid {grid-template-columns:repeat(2,1fr);}
            .page-title {font-size:22px;}
            section[data-testid="stSidebar"]:hover::before {left:266px;}
            section[data-testid="stSidebar"]:hover,
            section[data-testid="stSidebar"] > div:first-child {width:255px!important;min-width:255px!important;}
        }

footer {visibility: hidden;}
        .block-container {padding-bottom: 2rem;}

        .app-kicker {color:#8ea0b8; font-size:11px; font-weight:700; text-transform:uppercase; letter-spacing:1.4px; margin-bottom:2px;}
        .app-title {color:#f8fafc; font-size:28px; font-weight:800; letter-spacing:-.5px; margin-bottom:0;}
        .app-subtitle {color:#8796aa; font-size:13px; margin-top:3px;}

        .kanban-header {text-align:left; background:linear-gradient(135deg,#182235 0%,#111827 100%); color:#f8fafc!important; padding:13px 14px; border-radius:12px; font-weight:750; font-size:13px; border:1px solid #273449; margin-bottom:10px; box-shadow:0 8px 24px rgba(0,0,0,.14);}
        .kanban-count {color:#71819a!important; font-size:11px; font-weight:600;}
        .kanban-card {background:linear-gradient(145deg,#172131 0%,#111827 100%); border:1px solid #263449; border-radius:12px; padding:12px 46px 15px 14px; margin:0 0 12px 0; box-shadow:0 10px 26px rgba(0,0,0,.16);}
        .card-label {color:#73849a; font-size:10px; font-weight:700; text-transform:uppercase; letter-spacing:.7px;}
        .card-id {color:#60a5fa; font-size:12px; font-weight:800;}
        .card-driver {color:#f8fafc; font-size:13px; font-weight:800; margin:6px 0 6px;}
        .card-destination {color:#d6deea; font-size:11px; font-weight:600; margin-bottom:6px;}
        .card-meta {color:#a5b3c5; font-size:10px; line-height:1.6;}
        .card-divider {height:1px; background:#243145; margin:8px 0;}

        .metric-card {position:relative; overflow:hidden; background:linear-gradient(145deg,#172131 0%,#111827 100%); border:1px solid #263449; border-radius:12px; padding:13px 15px; min-height:92px; box-shadow:0 10px 28px rgba(0,0,0,.14);}
        .metric-card::after {content:""; position:absolute; right:-28px; top:-35px; width:90px; height:90px; border-radius:50%; background:rgba(255,255,255,.025);}
        .metric-title {color:#8998ac; font-size:10px; font-weight:800; letter-spacing:.8px;}
        .metric-value {color:#f8fafc; font-size:25px; font-weight:850; line-height:1.1; margin-top:7px;}
        .metric-subtitle {color:#718198; font-size:10px; margin-top:6px;}
        .metric-blue{border-left:3px solid #3b82f6;} .metric-green{border-left:3px solid #22c55e;} .metric-yellow{border-left:3px solid #f59e0b;} .metric-purple{border-left:3px solid #a855f7;} .metric-red{border-left:3px solid #ef4444;}

        .alert-box {padding:12px 15px; border-radius:11px; margin-bottom:10px; font-size:12px; font-weight:600;}
        .alert-red {background:rgba(239,68,68,.08); border:1px solid rgba(239,68,68,.28); color:#fca5a5;}
        .alert-yellow {background:rgba(245,158,11,.08); border:1px solid rgba(245,158,11,.28); color:#fcd34d;}
        .alert-green {background:rgba(34,197,94,.08); border:1px solid rgba(34,197,94,.25); color:#86efac;}

        .badge {display:inline-block; padding:3px 7px; border-radius:999px; font-size:9px; font-weight:800; letter-spacing:.25px; margin-right:4px; margin-bottom:4px;}
        .badge-red {background:rgba(239,68,68,.12); color:#fca5a5; border:1px solid rgba(239,68,68,.24);}
        .badge-yellow {background:rgba(245,158,11,.12); color:#fcd34d; border:1px solid rgba(245,158,11,.24);}
        .badge-green {background:rgba(34,197,94,.12); color:#86efac; border:1px solid rgba(34,197,94,.22);}
        .badge-blue {background:rgba(59,130,246,.12); color:#93c5fd; border:1px solid rgba(59,130,246,.22);}

        .section-title {color:#f1f5f9; font-size:18px; font-weight:800; margin:4px 0 14px;}
        .stSelectbox label,.stDateInput label,.stTextInput label,.stMultiSelect label,.stTextArea label {font-size:11px!important; color:#8ea0b8!important; font-weight:650!important;}
        div[data-testid="stDataFrame"] {border:1px solid #263449; border-radius:12px; overflow:hidden;}

        .chart-card {background:linear-gradient(145deg,#172131 0%,#111827 100%); border:1px solid #263449; border-radius:14px; padding:18px; margin:0 0 14px 0; box-shadow:0 10px 28px rgba(0,0,0,.14);}
        .chart-heading {display:flex; justify-content:space-between; align-items:flex-end; margin-bottom:16px; color:#f1f5f9; font-size:14px; font-weight:800;}
        .chart-heading small {color:#718198; font-size:10px; font-weight:600;}
        .donut-layout {display:flex; align-items:center; gap:30px; min-height:230px;}
        .donut {width:190px; height:190px; border-radius:50%; display:flex; align-items:center; justify-content:center; flex-shrink:0;}
        .donut-hole {width:116px; height:116px; border-radius:50%; background:#111827; display:flex; flex-direction:column; align-items:center; justify-content:center; box-shadow:0 0 0 1px #263449 inset;}
        .donut-hole strong {font-size:28px; line-height:1; color:#f8fafc;}
        .donut-hole span {font-size:10px; color:#718198; margin-top:5px;}
        .legend-list {flex:1;} .legend-row {display:grid; grid-template-columns:12px 1fr auto; gap:8px; align-items:center; padding:8px 0; color:#cbd5e1; font-size:11px; border-bottom:1px solid rgba(148,163,184,.08);}
        .legend-row b {color:#f8fafc; font-size:12px;} .legend-dot {width:8px; height:8px; border-radius:50%; display:block;}
        .bar-list {display:flex; flex-direction:column; gap:13px;}
        .bar-row {display:grid; grid-template-columns:180px 1fr 34px; gap:10px; align-items:center;}
        .bar-label {overflow:hidden; text-overflow:ellipsis; white-space:nowrap; color:#cbd5e1; font-size:11px;}
        .bar-track {height:9px; background:#202c3d; border-radius:999px; overflow:hidden;}
        .bar-fill {height:100%; border-radius:999px; background:linear-gradient(90deg,#3b82f6,#60a5fa); box-shadow:0 0 12px rgba(59,130,246,.18);}
        .bar-value {text-align:right; color:#f8fafc; font-size:12px; font-weight:800;}
        .chart-empty {padding:35px 10px; text-align:center; color:#718198; font-size:12px;}
        .evo-chart {height:220px; display:flex; align-items:flex-end; gap:10px; padding:12px 5px 0; border-bottom:1px solid #263449; overflow-x:auto;}
        .evo-item {height:100%; min-width:44px; display:flex; flex-direction:column; justify-content:flex-end; align-items:center; gap:6px;}
        .evo-value {color:#dbeafe; font-size:10px; font-weight:800;}
        .evo-bar {width:26px; min-height:8px; border-radius:7px 7px 2px 2px; background:linear-gradient(180deg,#60a5fa,#2563eb); box-shadow:0 4px 12px rgba(37,99,235,.18);}
        .evo-label {color:#718198; font-size:9px; white-space:nowrap;}

        @media (max-width:768px) {
            .block-container{padding-left:.65rem;padding-right:.65rem;padding-top:.6rem;}
            .app-title{font-size:22px;} .metric-card{min-height:96px;padding:13px;} .metric-value{font-size:24px;} .kanban-header{font-size:11px;padding:10px;}
        }
    
        /* =====================================================
           KANBAN DRAG & DROP + SETA DENTRO DO CARTÃO
           ===================================================== */

        .kanban-card[data-carga-id] {
            position: relative !important;
            cursor: grab !important;
            user-select: none;
            transition:
                border-color .15s ease,
                transform .15s ease,
                opacity .15s ease,
                box-shadow .15s ease !important;
        }

        .kanban-card[data-carga-id]:active {
            cursor: grabbing !important;
        }

        .kanban-card.dragging-card {
            opacity: .45 !important;
            transform: scale(.985) !important;
            border-color: #60a5fa !important;
            box-shadow: 0 0 0 1px rgba(96,165,250,.25) !important;
        }

        [data-testid="column"].kanban-dropzone {
            border-radius: 12px;
            transition: background .15s ease, box-shadow .15s ease;
        }

        [data-testid="column"].kanban-drop-active {
            background: rgba(59,130,246,.055) !important;
            box-shadow: inset 0 0 0 1px rgba(96,165,250,.24) !important;
        }

        [data-testid="column"].kanban-drop-active .kanban-header {
            color: #bfdbfe !important;
            border-bottom-color: #3b82f6 !important;
        }
/* Seta integrada ao próprio HTML do cartão */
        .kanban-card {
            position: relative !important;
            padding-right: 46px !important;
            padding-bottom: 16px !important;
        }

        .card-inline-arrow {
            position: absolute;
            right: 10px;
            bottom: 10px;
            width: 30px;
            height: 30px;
            border-radius: 7px;
            border: 1px solid #304158;
            background: #162131;
            color: #cbd5e1 !important;
            display: flex;
            align-items: center;
            justify-content: center;
            font-size: 19px;
            font-weight: 800;
            line-height: 1;
            text-decoration: none !important;
            cursor: pointer;
            z-index: 5;
            transition: .15s ease;
        }

        .card-inline-arrow:hover {
            background: #1a2a40;
            border-color: #3b82f6;
            color: #ffffff !important;
            transform: translateX(1px);
        }

        /* O cartão pode ser arrastado, mas a seta continua clicável */
        .card-inline-arrow {
            -webkit-user-drag: none;
        }

        .kanban-card.dragging-card {
            opacity: .42 !important;
            transform: scale(.985) !important;
        }

        .kanban-drop-active {
            background: rgba(59,130,246,.07) !important;
            box-shadow: inset 0 0 0 1px rgba(96,165,250,.35) !important;
        }

    </style>
    """,
    unsafe_allow_html=True
)


def render_html(html):
    """Renderiza HTML diretamente quando a versão do Streamlit oferece st.html."""
    if hasattr(st, "html"):
        st.html(html)
    else:
        st.markdown(html, unsafe_allow_html=True)




# ============================================================
# FIREBASE
# ============================================================

FIREBASE_PROJECT_ID = "logistica-d6c14"


# ============================================================
# FUNÇÕES AUXILIARES E CACHE
# ============================================================

FUSO_BRASIL = ZoneInfo("America/Sao_Paulo")

def agora_br():
    return datetime.datetime.now(FUSO_BRASIL)

def hoje_br():
    return agora_br().date()

def iso_agora_br():
    return agora_br().isoformat(timespec="seconds")

def converter_para_datetime(valor):
    if not valor:
        return None
    try:
        texto = str(valor).replace("Z", "+00:00")
        dt = datetime.datetime.fromisoformat(texto)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=FUSO_BRASIL)
        return dt.astimezone(FUSO_BRASIL)
    except Exception:
        return None

def formatar_data_br(data_str):
    if not data_str or str(data_str).lower() in ["nan", "none", ""]:
        return ""

    try:
        dt = datetime.datetime.strptime(
            str(data_str).split("T")[0],
            "%Y-%m-%d"
        )
        return dt.strftime("%d/%m/%Y")
    except Exception:
        return str(data_str)


def converter_para_data(data_str):
    """
    Converte datas armazenadas no Firebase para datetime.date.
    Retorna None quando não for possível converter.
    """
    if not data_str:
        return None

    try:
        return datetime.date.fromisoformat(
            str(data_str).split("T")[0]
        )
    except Exception:
        return None


@st.cache_data(ttl=60, show_spinner=False)
def carregar_colecao(colecao):
    url = (
        f"https://firestore.googleapis.com/v1/projects/"
        f"{FIREBASE_PROJECT_ID}/databases/(default)/documents/{colecao}"
    )

    try:
        response = requests.get(url, timeout=10)

        if response.status_code == 200:
            data = response.json()
            documentos = []

            for doc in data.get("documents", []):
                doc_id = doc["name"].split("/")[-1]
                fields = doc.get("fields", {})

                item = {"id": doc_id}

                for k, v in fields.items():

                    if "arrayValue" in v:
                        arr_vals = v["arrayValue"].get("values", [])

                        valores = []

                        for x in arr_vals:
                            if "stringValue" in x:
                                valores.append(x["stringValue"])
                            elif "integerValue" in x:
                                valores.append(int(x["integerValue"]))
                            elif "doubleValue" in x:
                                valores.append(float(x["doubleValue"]))
                            elif "booleanValue" in x:
                                valores.append(bool(x["booleanValue"]))

                        item[k] = valores

                    elif "stringValue" in v:
                        item[k] = v["stringValue"]

                    elif "integerValue" in v:
                        item[k] = int(v["integerValue"])

                    elif "doubleValue" in v:
                        item[k] = float(v["doubleValue"])

                    elif "booleanValue" in v:
                        item[k] = bool(v["booleanValue"])

                    elif "nullValue" in v:
                        item[k] = None

                    else:
                        item[k] = list(v.values())[0]

                documentos.append(item)

            return documentos

        else:
            st.warning(
                f"Não foi possível carregar '{colecao}'. "
                f"Firebase retornou HTTP {response.status_code}."
            )

    except requests.exceptions.Timeout:
        st.error(f"Tempo esgotado ao consultar a coleção '{colecao}'.")

    except requests.exceptions.RequestException as e:
        st.error(f"Erro de conexão com o Firebase: {e}")

    except Exception as e:
        st.error(f"Erro ao carregar '{colecao}': {e}")

    return []


def limpar_cache_firebase():
    """Limpa o cache para forçar nova busca no banco quando houver alterações."""
    carregar_colecao.clear()


def salvar_documento(colecao, doc_id, dados):
    url = (
        f"https://firestore.googleapis.com/v1/projects/"
        f"{FIREBASE_PROJECT_ID}/databases/(default)/documents/"
        f"{colecao}/{doc_id}"
    )

    fields = {}

    for k, v in dados.items():

        if isinstance(v, list):

            values = []

            for x in v:
                values.append({
                    "stringValue": str(x)
                })

            fields[k] = {
                "arrayValue": {
                    "values": values
                }
            }

        elif isinstance(v, bool):

            fields[k] = {
                "booleanValue": v
            }

        elif isinstance(v, int):

            fields[k] = {
                "integerValue": str(v)
            }

        elif isinstance(v, float):

            fields[k] = {
                "doubleValue": v
            }

        elif v is None:

            fields[k] = {
                "nullValue": None
            }

        else:

            fields[k] = {
                "stringValue": str(v)
            }

    try:

        response = requests.patch(
            url,
            json={"fields": fields},
            timeout=10
        )

        if response.status_code not in [200, 201]:

            st.error(
                f"Erro ao salvar no Firebase. "
                f"HTTP {response.status_code}"
            )

            return False

        limpar_cache_firebase()
        return True

    except requests.exceptions.Timeout:

        st.error("Tempo esgotado ao salvar no Firebase.")
        return False

    except requests.exceptions.RequestException as e:

        st.error(f"Erro de conexão com o Firebase: {e}")
        return False

    except Exception as e:

        st.error(f"Erro ao salvar no Firebase: {e}")
        return False


def atualizar_campos_documento(colecao, doc_id, campos):
    """
    Atualiza somente os campos informados no documento.
    Usado para movimentar o status automaticamente sem botão Salvar Status.
    """
    url = (
        f"https://firestore.googleapis.com/v1/projects/"
        f"{FIREBASE_PROJECT_ID}/databases/(default)/documents/"
        f"{colecao}/{doc_id}"
    )

    fields = {}

    for k, v in campos.items():

        if isinstance(v, list):
            fields[k] = {
                "arrayValue": {
                    "values": [
                        {"stringValue": str(x)}
                        for x in v
                    ]
                }
            }

        elif isinstance(v, bool):
            fields[k] = {"booleanValue": v}

        elif isinstance(v, int):
            fields[k] = {"integerValue": str(v)}

        elif isinstance(v, float):
            fields[k] = {"doubleValue": v}

        elif v is None:
            fields[k] = {"nullValue": None}

        else:
            fields[k] = {"stringValue": str(v)}

    params = [
        ("updateMask.fieldPaths", campo)
        for campo in campos.keys()
    ]

    try:
        response = requests.patch(
            url,
            params=params,
            json={"fields": fields},
            timeout=10
        )

        if response.status_code not in [200, 201]:
            st.error(
                f"Erro ao atualizar no Firebase. "
                f"HTTP {response.status_code}"
            )
            return False

        limpar_cache_firebase()
        return True

    except requests.exceptions.Timeout:
        st.error("Tempo esgotado ao atualizar o Firebase.")
        return False

    except requests.exceptions.RequestException as e:
        st.error(f"Erro de conexão com o Firebase: {e}")
        return False

    except Exception as e:
        st.error(f"Erro ao atualizar o Firebase: {e}")
        return False


def deletar_documento(colecao, doc_id):
    url = (
        f"https://firestore.googleapis.com/v1/projects/"
        f"{FIREBASE_PROJECT_ID}/databases/(default)/documents/"
        f"{colecao}/{doc_id}"
    )

    try:

        response = requests.delete(
            url,
            timeout=10
        )

        if response.status_code not in [200, 204]:

            st.error(
                f"Erro ao excluir documento. "
                f"HTTP {response.status_code}"
            )

            return False

        limpar_cache_firebase()
        return True

    except requests.exceptions.Timeout:

        st.error("Tempo esgotado ao excluir do Firebase.")
        return False

    except requests.exceptions.RequestException as e:

        st.error(f"Erro de conexão com o Firebase: {e}")
        return False

    except Exception as e:

        st.error(f"Erro ao excluir no Firebase: {e}")
        return False


def atualizar_dados():
    limpar_cache_firebase()
    st.session_state["cargas"] = carregar_colecao("cargas")
    st.session_state["motoristas"] = carregar_colecao("motoristas")
    st.session_state["ajudantes"] = carregar_colecao("ajudantes")
    st.session_state["veiculos"] = carregar_colecao("veiculos")
    st.session_state["ocorrencias"] = carregar_colecao("ocorrencias")

    st.session_state["ultima_atualizacao"] = agora_br().strftime(
        "%d/%m/%Y %H:%M:%S"
    )


def carga_atrasada(carga):
    status = carga.get("status", "")

    if status == "Entregue / Concluído":
        return False

    data_entrega = converter_para_data(
        carga.get("data_entrega")
    )

    if not data_entrega:
        return False

    return data_entrega < hoje_br()


def carga_saida_hoje(carga):
    data_saida = converter_para_data(
        carga.get("data_saida")
    )

    if not data_saida:
        return False

    return data_saida == hoje_br()


def carga_entrega_hoje(carga):
    data_entrega = converter_para_data(
        carga.get("data_entrega")
    )

    if not data_entrega:
        return False

    return data_entrega == hoje_br()


def dias_para_entrega(carga):
    data_entrega = converter_para_data(
        carga.get("data_entrega")
    )

    if not data_entrega:
        return None

    return (data_entrega - hoje_br()).days


def texto_prazo(carga):
    dias = dias_para_entrega(carga)

    if dias is None:
        return ""

    if carga_atrasada(carga):
        dias_atrasados = abs(dias)

        if dias_atrasados == 1:
            return "🔴 1 dia atrasada"

        return f"🔴 {dias_atrasados} dias atrasada"

    if dias == 0:
        return "🟠 Entrega hoje"

    if dias == 1:
        return "🟡 Entrega amanhã"

    if dias < 0:
        return "🔴 Atrasada"

    return f"🟢 {dias} dias para entrega"


def carga_em_risco(carga):
    if carga.get("status") == "Entregue / Concluído":
        return False
    dias = dias_para_entrega(carga)
    if dias is None or dias < 0:
        return False
    status = carga.get("status", "")
    # Entrega hoje e ainda não saiu, ou entrega amanhã e ainda aguarda carregamento.
    if dias == 0 and status in ["Aguardando Carregamento", "Carregado / No Pátio"]:
        return True
    if dias == 1 and status == "Aguardando Carregamento":
        return True
    return False

def classificacao_operacional(carga):
    if carga_atrasada(carga):
        return "🔴 Atrasada"
    if carga_em_risco(carga):
        return "🟠 Risco de atraso"
    if carga_entrega_hoje(carga):
        return "🟡 Entrega hoje"
    return "🟢 No prazo"

def intervalos_sobrepostos(inicio_a, fim_a, inicio_b, fim_b):
    return inicio_a <= fim_b and inicio_b <= fim_a

def conflitos_alocacao(cargas, motorista, veiculo, data_saida, data_entrega, ignorar_id=None):
    conflitos = []
    for c in cargas:
        if ignorar_id is not None and str(c.get("id")) == str(ignorar_id):
            continue
        if c.get("status") == "Entregue / Concluído":
            continue
        inicio = converter_para_data(c.get("data_saida"))
        fim = converter_para_data(c.get("data_entrega")) or inicio
        if not inicio or not fim:
            continue
        if intervalos_sobrepostos(data_saida, data_entrega, inicio, fim):
            if motorista and c.get("motorista") == motorista:
                conflitos.append(f"Motorista {motorista} já está no planejamento #{c.get('id')} ({formatar_data_br(c.get('data_saida'))} a {formatar_data_br(c.get('data_entrega'))}).")
            if veiculo and c.get("veiculo") == veiculo:
                conflitos.append(f"Veículo {veiculo} já está no planejamento #{c.get('id')} ({formatar_data_br(c.get('data_saida'))} a {formatar_data_br(c.get('data_entrega'))}).")
    return conflitos

def entrega_no_prazo(carga):
    previsto = converter_para_data(carga.get("data_entrega"))
    real_dt = converter_para_datetime(carga.get("data_hora_entrega_real"))
    if not previsto or not real_dt:
        return None
    return real_dt.date() <= previsto

def calcular_otd(cargas):
    avaliadas = [entrega_no_prazo(c) for c in cargas]
    avaliadas = [x for x in avaliadas if x is not None]
    if not avaliadas:
        return 0.0, 0
    return (sum(1 for x in avaliadas if x) / len(avaliadas)) * 100, len(avaliadas)

def preparar_dataframe(cargas_lista):
    df = pd.DataFrame(cargas_lista)

    if df.empty:
        return df

    if "ajudantes" in df.columns:
        df["ajudantes"] = df["ajudantes"].apply(
            lambda x: ", ".join(map(str, x))
            if isinstance(x, list)
            else str(x)
        )

    for col in [
        "data_carga",
        "data_saida",
        "data_entrega"
    ]:
        if col in df.columns:
            df[col] = df[col].apply(
                formatar_data_br
            )

    colunas_desejadas = [
        "id",
        "motorista",
        "veiculo",
        "destino",
        "observacoes",
        "ajudantes",
        "data_carga",
        "data_saida",
        "data_entrega",
        "data_hora_saida_real",
        "data_hora_entrega_real",
        "status"
    ]

    colunas_existentes = [
        col
        for col in colunas_desejadas
        if col in df.columns
    ]

    df = df[colunas_existentes]

    return df


# ============================================================
# EXCEL
# ============================================================

def gerar_excel_profissional(df):
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatorio de Cargas"

    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
    data_font = Font(name="Arial", size=9)
    center_alignment = Alignment(horizontal="center", vertical="center")
    left_alignment = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9")
    )

    mapa_headers = {
        "id": "ID Planejamento",
        "motorista": "Motorista",
        "veiculo": "Veículo",
        "destino": "Destino",
        "observacoes": "Observações",
        "ajudantes": "Ajudantes",
        "data_carga": "Data Carga",
        "data_saida": "Data Saída Prevista",
        "data_entrega": "Data Entrega Prevista",
        "data_hora_saida_real": "Saída Real",
        "data_hora_entrega_real": "Entrega Real",
        "status": "Status",
    }
    headers = [mapa_headers.get(col, col) for col in df.columns]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border

    ws.row_dimensions[1].height = 24

    for row_num, row_data in enumerate(df.values, 2):
        ws.row_dimensions[row_num].height = 20

        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = "" if str(value).lower() in ["nan", "none"] else value
            cell.font = data_font
            cell.border = thin_border

            if col_num in [1, 6, 7, 8]:
                cell.alignment = center_alignment
            else:
                cell.alignment = left_alignment

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 5, 15)

    wb.save(output)
    return output.getvalue()


# ============================================================
# PDF
# ============================================================

def gerar_pdf(df):
    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.add_page()
    pdf.set_font("Arial", "B", 14)
    pdf.cell(277, 8, txt="Relatório de Cargas", ln=True, align="C")

    pdf.set_font("Arial", "", 9)
    pdf.cell(277, 5, txt=f"Data de geração: {hoje_br().strftime('%d/%m/%Y')}", ln=True, align="C")
    pdf.ln(4)

    pdf.set_font("Arial", "B", 8)
    pdf.set_fill_color(47, 117, 181)
    pdf.set_text_color(255, 255, 255)

    colunas = [
        ("id", "ID", 20),
        ("motorista", "Motorista", 35),
        ("veiculo", "Veículo", 25),
        ("destino", "Destino", 45),
        ("data_saida", "Saída Prev.", 27),
        ("data_entrega", "Entrega Prev.", 27),
        ("status", "Status", 42),
        ("data_hora_entrega_real", "Entrega Real", 40),
    ]

    for _, nome, largura in colunas:
        pdf.cell(largura, 7, nome, 1, 0, "C", True)
    pdf.ln()
    pdf.set_font("Arial", "", 7)
    pdf.set_text_color(0, 0, 0)

    for _, row in df.iterrows():
        for chave, _, largura in colunas:
            valor = str(row.get(chave, "") or "")
            if chave in ["motorista", "destino", "status"]:
                limite = 26 if chave == "destino" else 22
                valor = valor[:limite]
                alinhamento = "L"
            else:
                valor = valor[:22]
                alinhamento = "C"
            pdf.cell(largura, 6, valor, 1, 0, alinhamento)
        pdf.ln()

    return pdf.output(dest="S").encode("latin-1")


# ============================================================
# CARREGAMENTO INICIAL
# ============================================================

if "cargas" not in st.session_state:
    st.session_state["cargas"] = carregar_colecao("cargas")

if "motoristas" not in st.session_state:
    st.session_state["motoristas"] = carregar_colecao("motoristas")

if "ajudantes" not in st.session_state:
    st.session_state["ajudantes"] = carregar_colecao("ajudantes")

if "veiculos" not in st.session_state:
    st.session_state["veiculos"] = carregar_colecao("veiculos")

if "ocorrencias" not in st.session_state:
    st.session_state["ocorrencias"] = carregar_colecao("ocorrencias")


# ============================================================
# LISTAS DA EQUIPE
# ============================================================

cargas_lista = st.session_state["cargas"]

motoristas_lista = [
    m.get("nome", "")
    for m in st.session_state["motoristas"]
    if m.get("nome")
]

ajudantes_lista = [
    a.get("nome", "")
    for a in st.session_state["ajudantes"]
    if a.get("nome")
]

if not motoristas_lista:
    motoristas_lista = ["Carlos Silva", "João Pereira", "Maurício", "Cícero Taveira"]

if not ajudantes_lista:
    ajudantes_lista = ["Pedrinho", "Lucas Souza"]

veiculos_lista = [
    v.get("placa", "")
    for v in st.session_state["veiculos"]
    if v.get("placa") and v.get("ativo", True)
]


# ============================================================
# RESUMO OPERACIONAL GLOBAL
# ============================================================

hoje = hoje_br()
total_cargas = len(cargas_lista)
total_aguardando = sum(1 for c in cargas_lista if c.get("status") == "Aguardando Carregamento")
total_patio = sum(1 for c in cargas_lista if c.get("status") == "Carregado / No Pátio")
total_transito = sum(1 for c in cargas_lista if c.get("status") == "Em Trânsito / Viagem Iniciada")
total_entregues = sum(1 for c in cargas_lista if c.get("status") == "Entregue / Concluído")
total_atrasadas = sum(1 for c in cargas_lista if carga_atrasada(c))
total_saida_hoje = sum(1 for c in cargas_lista if carga_saida_hoje(c))
total_entrega_hoje = sum(1 for c in cargas_lista if carga_entrega_hoje(c))
total_risco = sum(1 for c in cargas_lista if carga_em_risco(c))
ocorrencias_abertas = [
    o for o in st.session_state.get("ocorrencias", [])
    if o.get("status", "Aberta") == "Aberta"
]


# ============================================================
# MENU LATERAL FLUTUANTE
# ============================================================

with st.sidebar:
    st.markdown("""
        <div style="padding:4px 4px 15px;white-space:nowrap;">
            <div style="font-size:18px;color:#f8fafc;font-weight:900;">
                🚚 Gestão de Cargas
            </div>
            <div style="font-size:9px;color:#718198;font-weight:800;letter-spacing:1.15px;margin:5px 0 14px 30px;">
                LOGÍSTICA
            </div>
        </div>
    """, unsafe_allow_html=True)

    st.caption("OPERAÇÃO")
    menu = st.radio(
        "Navegação",
        [
            "🏠 Visão Geral",
            "📋 Torre de Controle",
            "➕ Nova Carga",
            "🚨 Ocorrências",
            "📈 Relatórios",
            "👥 Cadastros",
        ],
        label_visibility="collapsed",
        key="menu_principal",
    )

    st.markdown("---")
    if st.button("↻ Atualizar dados", use_container_width=True):
        atualizar_dados()
        st.rerun()

    if "ultima_atualizacao" in st.session_state:
        st.caption("Última atualização")
        st.caption(st.session_state["ultima_atualizacao"])


# ============================================================
# CABEÇALHO DA ÁREA DE TRABALHO
# ============================================================

cabecalhos_menu = {
    "🏠 Visão Geral": ("Visão Geral", "Panorama da operação e prioridades do dia."),
    "📋 Torre de Controle": ("Torre de Controle", "Acompanhe e movimente as cargas da operação."),
    "➕ Nova Carga": ("Nova Carga", "Cadastre e programe um novo planejamento de transporte."),
    "🚨 Ocorrências": ("Central de Ocorrências", "Registre e acompanhe desvios que exigem atenção."),
    "📈 Relatórios": ("Relatórios", "Analise desempenho, entregas e indicadores operacionais."),
    "👥 Cadastros": ("Equipe e Frota", "Gerencie motoristas, ajudantes e veículos."),
}

titulo_pagina, subtitulo_pagina = cabecalhos_menu[menu]
st.markdown(
    f"""
    <div class="page-head">
        <div>
            <div class="page-title">{titulo_pagina}</div>
            <div class="page-subtitle">{subtitulo_pagina}</div>
        </div>
        <div class="date-chip">📅 {hoje_br().strftime('%d/%m/%Y')}</div>
    </div>
    """,
    unsafe_allow_html=True,
)


# ============================================================
# 0. VISÃO GERAL
# ============================================================

if menu == "🏠 Visão Geral":

    st.markdown(
        f"""
        <div class="kpi-grid">
            <div class="kpi-card"><div class="kpi-icon">📋</div><div class="kpi-title">Programadas</div><div class="kpi-value">{total_cargas}</div><div class="kpi-sub">Total de cargas</div></div>
            <div class="kpi-card"><div class="kpi-icon" style="color:#fbbf24;">◷</div><div class="kpi-title">Aguardando</div><div class="kpi-value">{total_aguardando}</div><div class="kpi-sub">Carregamento</div></div>
            <div class="kpi-card"><div class="kpi-icon" style="color:#c084fc;">⌂</div><div class="kpi-title">No pátio</div><div class="kpi-value">{total_patio}</div><div class="kpi-sub">Carregadas</div></div>
            <div class="kpi-card"><div class="kpi-icon">🚚</div><div class="kpi-title">Em trânsito</div><div class="kpi-value">{total_transito}</div><div class="kpi-sub">Viagens</div></div>
            <div class="kpi-card"><div class="kpi-icon" style="color:#4ade80;">✓</div><div class="kpi-title">Entregues</div><div class="kpi-value">{total_entregues}</div><div class="kpi-sub">Concluídas</div></div>
            <div class="kpi-card"><div class="kpi-icon" style="color:#f59e0b;">⚠</div><div class="kpi-title">Em risco</div><div class="kpi-value">{total_risco}</div><div class="kpi-sub">Atenção</div></div>
            <div class="kpi-card"><div class="kpi-icon" style="color:#fb7185;">!</div><div class="kpi-title">Atrasadas</div><div class="kpi-value">{total_atrasadas}</div><div class="kpi-sub">Fora do prazo</div></div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    itens_atencao = []
    for carga in cargas_lista:
        if carga_atrasada(carga):
            itens_atencao.append((0, carga, f"Entrega atrasada {abs(dias_para_entrega(carga) or 0)} dia(s)", "attention-danger"))
        elif carga_em_risco(carga):
            mensagem = "Entrega hoje e operação ainda não concluída" if carga_entrega_hoje(carga) else "Risco de atraso"
            itens_atencao.append((1, carga, mensagem, "attention-warning"))
        elif carga_saida_hoje(carga) and carga.get("status") == "Aguardando Carregamento":
            itens_atencao.append((2, carga, "Saída hoje, carregamento não iniciado", "attention-warning"))

    itens_atencao.sort(key=lambda x: x[0])

    linhas_atencao = ""
    for _, carga, mensagem, classe in itens_atencao[:6]:
        linhas_atencao += f"""
        <div class="attention-row">
            <div class="attention-id">#{carga.get('id','—')}</div>
            <div class="{classe}">{mensagem}</div>
            <div>👤 {carga.get('motorista','') or 'Sem motorista'}</div>
            <div class="hide-small">🚛 {carga.get('veiculo','') or 'Sem veículo'}</div>
            <div class="hide-small">📍 {carga.get('destino','') or 'Sem destino'}</div>
        </div>
        """

    for ocorrencia in ocorrencias_abertas[:3]:
        descricao = ocorrencia.get("descricao") or ocorrencia.get("tipo") or "Ocorrência aberta"
        linhas_atencao += f"""
        <div class="attention-row">
            <div class="attention-id">#{ocorrencia.get('carga_id','—')}</div>
            <div class="attention-danger">🚨 {descricao}</div>
            <div>Ocorrência aberta</div>
            <div class="hide-small">—</div>
            <div class="hide-small">Requer tratamento</div>
        </div>
        """

    if not linhas_atencao:
        linhas_atencao = '<div class="attention-empty">✓ Nenhuma carga crítica ou ocorrência aberta neste momento.</div>'

    st.markdown(
        f"""
        <div class="attention-panel">
            <div class="attention-title">⚠ REQUER ATENÇÃO</div>
            {linhas_atencao}
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.markdown('<div class="section-caption">Situação das Cargas</div>', unsafe_allow_html=True)

    status_preview = [
        ("Aguardando Carregamento", "Aguardando Carregamento", "#f59e0b"),
        ("Carregado / No Pátio", "No Pátio", "#a855f7"),
        ("Em Trânsito / Viagem Iniciada", "Em Trânsito", "#3b82f6"),
        ("Entregue / Concluído", "Entregues", "#22c55e"),
    ]

    preview_cols = st.columns(4)
    for idx, (status_real, titulo_status, cor) in enumerate(status_preview):
        cargas_status = [c for c in cargas_lista if c.get("status") == status_real]
        with preview_cols[idx]:
            st.markdown(
                f"""
                <div class="kanban-header" style="border-left:2px solid {cor}!important;">
                    {titulo_status}
                    <span class="kanban-count">{len(cargas_status)}</span>
                </div>
                """,
                unsafe_allow_html=True,
            )
            for carga in cargas_status[:3]:
                if carga_atrasada(carga):
                    badge = '<span class="badge badge-red">ATRASADA</span>'
                elif carga_em_risco(carga):
                    badge = '<span class="badge badge-yellow">RISCO</span>'
                elif status_real == "Entregue / Concluído":
                    badge = '<span class="badge badge-green">✓</span>'
                else:
                    badge = ""

                entrega_label = "Entregue" if status_real == "Entregue / Concluído" else "Entrega"
                entrega_valor = formatar_data_br(carga.get("data_entrega")) or "—"

                st.markdown(
                    f"""
                    <div class="kanban-card">
                        <div class="card-topline">
                            <div class="card-id">#{carga.get('id','')}</div>
                            <div>{badge}</div>
                        </div>
                        <div class="card-destination">{carga.get('destino','') or 'Sem destino'}</div>
                        <div class="card-meta">👤 {carga.get('motorista','') or 'Sem motorista'} &nbsp; 🚛 {carga.get('veiculo','') or 'Sem veículo'}</div>
                        <div class="card-deadline">{entrega_label}: <strong>{entrega_valor}</strong></div>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )

            if len(cargas_status) > 3:
                st.caption(f"+ {len(cargas_status) - 3} outra(s) carga(s)")


# ============================================================
# 1. PAINEL KANBAN
# ============================================================

elif menu == "📋 Torre de Controle":

    # FILTROS
    st.markdown('<div class="section-caption">Torre de Controle</div>', unsafe_allow_html=True)
    col_f1, col_f2 = st.columns([2, 3])

    with col_f1:
        motoristas_filtro_opcoes = ["Todos os Motoristas"] + motoristas_lista
        motorista_selecionado = st.selectbox("Filtrar por Motorista", motoristas_filtro_opcoes)

    with col_f2:
        pesquisa = st.text_input("🔎 Pesquisar", placeholder="ID, motorista, destino ou observação...")

    cargas_filtradas_periodo = list(cargas_lista)

    if motorista_selecionado != "Todos os Motoristas":
        cargas_filtradas_periodo = [c for c in cargas_filtradas_periodo if c.get("motorista") == motorista_selecionado]

    if pesquisa:
        termo = pesquisa.lower().strip()
        cargas_filtradas_periodo = [
            c for c in cargas_filtradas_periodo
            if (
                termo in str(c.get("id", "")).lower()
                or termo in str(c.get("motorista", "")).lower()
                or termo in str(c.get("destino", "")).lower()
                or termo in str(c.get("observacoes", "")).lower()
                or termo in str(c.get("status", "")).lower()
            )
        ]

    st.caption(f"🔎 {len(cargas_filtradas_periodo)} carga(s) encontrada(s) com os filtros atuais.")

    colunas_status = [
        "Aguardando Carregamento",
        "Carregado / No Pátio",
        "Em Trânsito / Viagem Iniciada",
        "Entregue / Concluído",
    ]


    # ============================================================
    # MOVIMENTAÇÃO POR ARRASTAR E SOLTAR
    # ============================================================
    carga_id_drag = st.query_params.get("mover_carga")
    novo_status_drag = st.query_params.get("mover_status")

    if carga_id_drag and novo_status_drag:
        carga_drag = next(
            (c for c in cargas_lista if str(c.get("id")) == str(carga_id_drag)),
            None,
        )

        if (
            carga_drag
            and novo_status_drag in colunas_status
            and carga_drag.get("status") != novo_status_drag
        ):
            campos_status_drag = {"status": novo_status_drag}

            if (
                novo_status_drag == "Em Trânsito / Viagem Iniciada"
                and not carga_drag.get("data_hora_saida_real")
            ):
                campos_status_drag["data_hora_saida_real"] = iso_agora_br()

            if novo_status_drag == "Entregue / Concluído":
                campos_status_drag["data_conclusao"] = str(hoje_br())
                if not carga_drag.get("data_hora_entrega_real"):
                    campos_status_drag["data_hora_entrega_real"] = iso_agora_br()

            sucesso_drag = atualizar_campos_documento(
                "cargas",
                carga_id_drag,
                campos_status_drag,
            )

            if sucesso_drag:
                carga_drag.update(campos_status_drag)

                # Atualiza também a lista da sessão para a mudança aparecer
                # imediatamente na coluna correta.
                for item in st.session_state["cargas"]:
                    if str(item.get("id")) == str(carga_id_drag):
                        item.update(campos_status_drag)
                        break

                st.session_state[f"editando_{carga_id_drag}"] = False

                # Só limpa a URL DEPOIS de salvar no Firebase.
                try:
                    st.query_params.clear()
                except Exception:
                    pass

                st.toast(
                    f"Carga #{carga_id_drag} movida para {novo_status_drag}.",
                    icon="✅",
                )
                st.rerun()

        else:
            # Se a movimentação não é válida ou é para a mesma coluna,
            # limpa os parâmetros sem alterar a carga.
            try:
                st.query_params.clear()
            except Exception:
                pass

    st.caption("↔️ Clique, segure e arraste a carga para outra coluna.")

    carga_acao_id = st.query_params.get("acao_carga")
    if carga_acao_id:
        st.session_state["_acao_carga_id"] = str(carga_acao_id)
        try:
            del st.query_params["acao_carga"]
        except Exception:
            pass


    paleta_cores = ["#58a6ff", "#3fb950", "#d29922", "#bc8cff", "#f85149", "#39c5bb", "#f0883e", "#db61a2"]
    mapa_cores = {mot: paleta_cores[i % len(paleta_cores)] for i, mot in enumerate(motoristas_lista)}

    # KANBAN
    cols = st.columns(len(colunas_status))

    for idx, status in enumerate(colunas_status):
        with cols[idx]:
            quantidade_status = sum(1 for c in cargas_filtradas_periodo if c.get("status") == status)
            st.markdown(f"""
                <div class='kanban-header' data-kanban-status="{status}">
                    <div>{status}</div>
                    <div class="kanban-count">{quantidade_status} carga(s)</div>
                </div>
                """, unsafe_allow_html=True)

            cargas_status_filtradas = [c for c in cargas_filtradas_periodo if c.get("status") == status]

            for carga in cargas_status_filtradas:
                carga_id = carga.get("id")
                motorista_atual = carga.get("motorista", "")
                veiculo_atual = carga.get("veiculo", "")
                cor_motorista = mapa_cores.get(motorista_atual, "#8b949e")
                saida_br = formatar_data_br(carga.get("data_saida"))
                entrega_br = formatar_data_br(carga.get("data_entrega"))
                ajudantes = carga.get("ajudantes", [])
                
                if isinstance(ajudantes, list):
                    ajudantes_texto = ", ".join(map(str, ajudantes))
                else:
                    ajudantes_texto = str(ajudantes)

                atrasada = carga_atrasada(carga)
                saida_hoje = carga_saida_hoje(carga)
                entrega_hoje = carga_entrega_hoje(carga)
                prazo_texto = texto_prazo(carga)

                with st.container():
                    cor_borda = "#f85149" if atrasada else cor_motorista
                    badges = ""

                    if atrasada:
                        badges += '<span class="badge badge-red">🔴 ATRASADA</span>'
                    elif entrega_hoje:
                        badges += '<span class="badge badge-yellow">📦 ENTREGA HOJE</span>'
                    elif saida_hoje:
                        badges += '<span class="badge badge-yellow">🚚 SAI HOJE</span>'
                    else:
                        badges += '<span class="badge badge-green">✓ NO PRAZO</span>'

                    if prazo_texto:
                        badges += f'<span class="badge badge-blue">{prazo_texto}</span>'

                    observacoes = str(carga.get("observacoes", "")).strip()

                    if atrasada:
                        badge_card = '<span class="badge badge-red">ATRASADA</span>'
                    elif carga_em_risco(carga):
                        badge_card = '<span class="badge badge-yellow">RISCO</span>'
                    elif status == "Entregue / Concluído":
                        badge_card = '<span class="badge badge-green">✓ ENTREGUE</span>'
                    elif entrega_hoje:
                        badge_card = '<span class="badge badge-yellow">ENTREGA HOJE</span>'
                    else:
                        badge_card = ""

                    render_html(f"""
                        <div
                            class="kanban-card"
                            data-carga-id="{carga_id}"
                            data-current-status="{status}"
                            draggable="true"
                            style="border-left:2px solid {cor_borda};"
                        >
                            <div class="card-topline">
                                <div class="card-id">#{carga_id}</div>
                                <div>{badge_card}</div>
                            </div>
                            <div class="card-destination">{carga.get('destino', '') or 'Sem destino'}</div>
                            <div class="card-meta">👤 {motorista_atual or 'Sem motorista'} &nbsp; 🚛 {veiculo_atual or 'Sem veículo'}</div>
                            <div class="card-deadline">Entrega: <strong>{entrega_br or '—'}</strong></div>
                            <a
                                class="card-inline-arrow"
                                href="?acao_carga={carga_id}"
                                draggable="false"
                                title="Abrir ações da carga"
                                onclick="event.stopPropagation();"
                            >›</a>
                        </div>
                    """)

                    # ==================================================
                    # AÇÕES DA CARGA SELECIONADA PELA SETA DO CARTÃO
                    # ==================================================
                    if st.session_state.get("_acao_carga_id") == str(carga_id):
                        with st.container(border=True):
                            ac1, ac2, ac3 = st.columns([3, 1, 1])

                            with ac1:
                                st.caption(f"Ações do planejamento #{carga_id}")

                            with ac2:
                                if st.button(
                                    "✏️ Editar",
                                    key=f"btn_edit_{carga_id}",
                                    use_container_width=True,
                                ):
                                    st.session_state[f"editando_{carga_id}"] = not st.session_state.get(
                                        f"editando_{carga_id}",
                                        False,
                                    )
                                    st.session_state["_acao_carga_id"] = None
                                    st.rerun()

                            with ac3:
                                if st.button(
                                    "🗑️ Excluir",
                                    key=f"btn_del_{carga_id}",
                                    use_container_width=True,
                                ):
                                    st.session_state[f"confirmar_exclusao_{carga_id}"] = True
                                    st.session_state["_acao_carga_id"] = None
                                    st.rerun()

                            if st.button(
                                "Fechar",
                                key=f"fechar_acoes_{carga_id}",
                                use_container_width=True,
                            ):
                                st.session_state["_acao_carga_id"] = None
                                st.rerun()

                    # Confirmação de exclusão
                    if st.session_state.get(f"confirmar_exclusao_{carga_id}", False):
                        st.warning(f"Tem certeza que deseja excluir a carga/planejamento {carga_id}?")
                        cx1, cx2 = st.columns(2)
                        with cx1:
                            if st.button("Sim, excluir", key=f"confirm_del_{carga_id}", type="primary"):
                                sucesso = deletar_documento("cargas", carga_id)
                                if sucesso:
                                    st.session_state["cargas"] = [c for c in cargas_lista if c.get("id") != carga_id]
                                    st.session_state[f"confirmar_exclusao_{carga_id}"] = False
                                    st.success("Carga excluída.")
                                    st.rerun()
                        with cx2:
                            if st.button("Cancelar", key=f"cancel_del_{carga_id}"):
                                st.session_state[f"confirmar_exclusao_{carga_id}"] = False
                                st.rerun()


                    # ==================================================
                    # FORMULÁRIO DE EDIÇÃO
                    # ==================================================
                    if st.session_state.get(f"editando_{carga_id}", False):
                        with st.form(key=f"form_edit_{carga_id}"):
                            st.markdown(f"**✏️ Editando Planejamento #{carga_id}**")
                            
                            mot_idx = motoristas_lista.index(carga.get("motorista")) if carga.get("motorista") in motoristas_lista else 0
                            novo_mot = st.selectbox("Motorista", motoristas_lista if motoristas_lista else [""], index=mot_idx)
                            veic_opcoes = ["Não definido"] + veiculos_lista
                            veic_atual = carga.get("veiculo", "") or "Não definido"
                            veic_idx = veic_opcoes.index(veic_atual) if veic_atual in veic_opcoes else 0
                            novo_veiculo = st.selectbox("Veículo", veic_opcoes, index=veic_idx, key=f"veiculo_{carga_id}")
                            novo_dest = st.text_input("Destino", value=carga.get("destino", ""))
                            novo_obs = st.text_area("Observações / Rota", value=carga.get("observacoes", ""))
                            
                            ajudantes_existentes = carga.get("ajudantes", [])
                            if not isinstance(ajudantes_existentes, list):
                                ajudantes_existentes = []

                            ajudantes_editados = st.multiselect(
                                "Ajudantes",
                                ajudantes_lista,
                                default=[a for a in ajudantes_existentes if a in ajudantes_lista]
                            )

                            dt_saida_val = converter_para_data(carga.get("data_saida")) or hoje_br()
                            dt_ent_val = converter_para_data(carga.get("data_entrega")) or hoje_br()

                            nova_saida = st.date_input("Data Saída", value=dt_saida_val, key=f"saida_{carga_id}")
                            nova_entrega = st.date_input("Data Entrega", value=dt_ent_val, key=f"entrega_{carga_id}")

                            salvar_edicao = st.form_submit_button("💾 Salvar Alterações")

                            if salvar_edicao:
                                novo_veiculo_valor = "" if novo_veiculo == "Não definido" else novo_veiculo
                                conflitos = conflitos_alocacao(cargas_lista, novo_mot, novo_veiculo_valor, nova_saida, nova_entrega, ignorar_id=carga_id)
                                if nova_entrega < nova_saida:
                                    st.error("A data de entrega não pode ser anterior à saída.")
                                    st.stop()
                                if conflitos:
                                    st.error("Conflito de programação encontrado:")
                                    for conflito in conflitos:
                                        st.warning(conflito)
                                    st.stop()
                                carga["motorista"] = novo_mot
                                carga["veiculo"] = novo_veiculo_valor
                                carga["destino"] = novo_dest
                                carga["observacoes"] = novo_obs
                                carga["ajudantes"] = ajudantes_editados
                                carga["data_saida"] = str(nova_saida)
                                carga["data_entrega"] = str(nova_entrega)

                                sucesso = salvar_documento("cargas", carga_id, carga)

                                if sucesso:
                                    st.session_state[f"editando_{carga_id}"] = False
                                    st.success("Carga atualizada com sucesso!")
                                    st.rerun()



    # Ativa drag-and-drop no DOM da página principal.
    components.html(
        """
        <script>
        (() => {
            const doc = window.parent.document;
            const win = window.parent;

            function limpar(v) {
                return (v || "").trim();
            }

            function moverCarga(cargaId, status) {
                if (!cargaId || !status) return;

                const url = new URL(win.location.href);
                url.searchParams.delete("acao_carga");
                url.searchParams.set("mover_carga", cargaId);
                url.searchParams.set("mover_status", status);
                win.location.href = url.toString();
            }

            function preparar() {
                const headers = [...doc.querySelectorAll(
                    '.kanban-header[data-kanban-status]'
                )];

                const cards = [...doc.querySelectorAll(
                    '.kanban-card[data-carga-id]'
                )];

                if (!headers.length || !cards.length) return;

                const columns = [];

                headers.forEach((header) => {
                    const col = header.closest('[data-testid="column"]');
                    if (!col) return;

                    const status = limpar(header.dataset.kanbanStatus);
                    if (!status) return;

                    col.dataset.kanbanStatus = status;
                    col.classList.add("kanban-dropzone");
                    columns.push(col);

                    if (col.dataset.dndReady === "1") return;
                    col.dataset.dndReady = "1";

                    // Capture=true ajuda quando o drop acontece sobre widgets filhos.
                    col.addEventListener("dragover", (ev) => {
                        ev.preventDefault();
                        ev.stopPropagation();
                        if (ev.dataTransfer) {
                            ev.dataTransfer.dropEffect = "move";
                        }
                        col.classList.add("kanban-drop-active");
                    }, true);

                    col.addEventListener("dragenter", (ev) => {
                        ev.preventDefault();
                        ev.stopPropagation();
                        col.classList.add("kanban-drop-active");
                    }, true);

                    col.addEventListener("dragleave", (ev) => {
                        const rect = col.getBoundingClientRect();
                        const x = ev.clientX;
                        const y = ev.clientY;

                        const fora =
                            x <= rect.left ||
                            x >= rect.right ||
                            y <= rect.top ||
                            y >= rect.bottom;

                        if (fora) {
                            col.classList.remove("kanban-drop-active");
                        }
                    }, true);

                    col.addEventListener("drop", (ev) => {
                        ev.preventDefault();
                        ev.stopPropagation();

                        const cargaId =
                            ev.dataTransfer?.getData("text/plain") || "";

                        const origem =
                            ev.dataTransfer?.getData("application/x-origem") || "";

                        const destino = limpar(col.dataset.kanbanStatus);

                        doc.querySelectorAll(".kanban-drop-active").forEach((el) => {
                            el.classList.remove("kanban-drop-active");
                        });

                        if (
                            cargaId &&
                            destino &&
                            limpar(origem) !== destino
                        ) {
                            moverCarga(cargaId, destino);
                        }
                    }, true);
                });

                cards.forEach((card) => {
                    card.setAttribute("draggable", "true");

                    if (card.dataset.dndReady === "1") return;
                    card.dataset.dndReady = "1";

                    card.addEventListener("dragstart", (ev) => {
                        // Não começa drag se iniciou pela seta.
                        if (ev.target?.closest?.(".card-inline-arrow")) {
                            ev.preventDefault();
                            return;
                        }

                        const id = card.dataset.cargaId || "";
                        const origem = limpar(card.dataset.currentStatus);

                        if (!id || !ev.dataTransfer) {
                            ev.preventDefault();
                            return;
                        }

                        card.classList.add("dragging-card");
                        ev.dataTransfer.effectAllowed = "move";
                        ev.dataTransfer.setData("text/plain", id);
                        ev.dataTransfer.setData("application/x-origem", origem);
                    });

                    card.addEventListener("dragend", () => {
                        card.classList.remove("dragging-card");

                        doc.querySelectorAll(".kanban-drop-active").forEach((el) => {
                            el.classList.remove("kanban-drop-active");
                        });
                    });
                });
            }

            let loops = 0;
            const timer = setInterval(() => {
                loops += 1;
                preparar();
                if (loops >= 50) clearInterval(timer);
            }, 100);

            const observer = new MutationObserver(preparar);
            observer.observe(doc.body, {childList:true, subtree:true});

            setTimeout(() => observer.disconnect(), 20000);
        })();
        </script>
        """,
        height=0,
        width=0,
    )


# ============================================================
# 2. NOVA CARGA
# ============================================================

elif menu == "➕ Nova Carga":

    st.info("📌 O número do planejamento é fornecido pelo seu sistema de montagem de cargas. Digite exatamente o número recebido.")

    with st.form("form_nova_carga"):
        col_id_manual, _ = st.columns([2, 2])

        with col_id_manual:
            id_planejamento = st.text_input("Número do Planejamento / ID da Carga", placeholder="Ex: 1042")

        col1, col2 = st.columns(2)

        with col1:
            motorista = st.selectbox("Motorista Responsável", motoristas_lista if motoristas_lista else ["Nenhum cadastrado"])
            veiculo = st.selectbox("Veículo", ["Não definido"] + veiculos_lista)
            destino = st.text_input("Região / Cidades de Destino", placeholder="Ex: Uberaba, Araxá (Múltiplas entregas)")
            observacoes = st.text_area("Observações / Rota", placeholder="Ex: Carga com entregas em lojas diferentes")

        with col2:
            ajudantes = st.multiselect("Ajudantes da Viagem", ajudantes_lista)
            data_carga = st.date_input("Data do Carregamento")
            data_saida = st.date_input("Data de Saída")
            data_entrega = st.date_input("Data Prevista de Entrega")

        status_inicial = st.selectbox(
            "Status Inicial",
            [
                "Aguardando Carregamento",
                "Carregado / No Pátio",
                "Em Trânsito / Viagem Iniciada",
                "Entregue / Concluído",
            ],
        )

        submit = st.form_submit_button("💾 Salvar e Agendar Carga", type="primary")

        if submit:
            id_planejamento = str(id_planejamento).strip()

            if id_planejamento and destino and motorista:
                ids_existentes = [str(c.get("id")) for c in cargas_lista]
                veiculo_valor = "" if veiculo == "Não definido" else veiculo
                conflitos = conflitos_alocacao(cargas_lista, motorista, veiculo_valor, data_saida, data_entrega)

                if id_planejamento in ids_existentes:
                    st.error(f"Já existe uma carga cadastrada com o ID/Planejamento '{id_planejamento}'.")
                elif data_entrega < data_saida:
                    st.error("A data prevista de entrega não pode ser anterior à data de saída.")
                elif conflitos:
                    st.error("Conflito de programação encontrado:")
                    for conflito in conflitos:
                        st.warning(conflito)
                else:
                    nova_carga = {
                        "id": id_planejamento,
                        "motorista": motorista,
                        "veiculo": veiculo_valor,
                        "destino": destino,
                        "observacoes": observacoes,
                        "ajudantes": ajudantes,
                        "data_carga": str(data_carga),
                        "data_saida": str(data_saida),
                        "data_entrega": str(data_entrega),
                        "status": status_inicial,
                    }

                    sucesso = salvar_documento("cargas", id_planejamento, nova_carga)

                    if sucesso:
                        st.session_state["cargas"].append(nova_carga)
                        st.success(f"✅ Carga/Planejamento {id_planejamento} cadastrada com sucesso!")
                        st.rerun()
            else:
                st.error("Preencha o Número do Planejamento, o Motorista e a Região de Destino.")


# ============================================================
# 3. CADASTROS
# ============================================================

elif menu == "👥 Cadastros":

    col1, col2 = st.columns(2)

    with col1:
        st.markdown("### 🚚 Motoristas")

        with st.form("form_cad_mot", clear_on_submit=True):
            novo_mot = st.text_input("Adicionar novo motorista")
            cad_mot_btn = st.form_submit_button("Cadastrar Motorista")

            if cad_mot_btn and novo_mot:
                novo_mot = novo_mot.strip()
                nomes_existentes = [str(m.get("nome", "")).lower() for m in st.session_state["motoristas"]]

                if novo_mot.lower() in nomes_existentes:
                    st.error("Este motorista já está cadastrado.")
                else:
                    doc_id = f"mot_{int(agora_br().timestamp() * 1000)}"
                    dados_mot = {"id": doc_id, "nome": novo_mot}
                    sucesso = salvar_documento("motoristas", doc_id, dados_mot)

                    if sucesso:
                        st.session_state["motoristas"].append(dados_mot)
                        st.success(f"Motorista {novo_mot} adicionado!")
                        st.rerun()

        st.markdown("---")
        st.write("**Motoristas Atuais:**")

        if not st.session_state["motoristas"]:
            st.info("Nenhum motorista cadastrado.")

        for m_obj in st.session_state["motoristas"]:
            m_nome = m_obj.get("nome", "")
            m_id = m_obj.get("id", m_nome)
            c_mot1, c_mot2 = st.columns([4, 2])

            c_mot1.write(f"🚚 {m_nome}")

            if c_mot2.button("Excluir", key=f"del_mot_{m_id}"):
                st.session_state[f"confirmar_mot_{m_id}"] = True

            if st.session_state.get(f"confirmar_mot_{m_id}", False):
                st.warning(f"Excluir o motorista '{m_nome}'?")
                cm1, cm2 = st.columns(2)

                with cm1:
                    if st.button("Confirmar", key=f"conf_mot_{m_id}"):
                        sucesso = deletar_documento("motoristas", m_id)
                        if sucesso:
                            st.session_state["motoristas"] = [m for m in st.session_state["motoristas"] if m.get("id") != m_id]
                            st.rerun()

                with cm2:
                    if st.button("Cancelar", key=f"canc_mot_{m_id}"):
                        st.session_state[f"confirmar_mot_{m_id}"] = False
                        st.rerun()

    with col2:
        st.markdown("### 👥 Ajudantes")

        with st.form("form_cad_aju", clear_on_submit=True):
            novo_aju = st.text_input("Adicionar novo ajudante")
            cad_aju_btn = st.form_submit_button("Cadastrar Ajudante")

            if cad_aju_btn and novo_aju:
                novo_aju = novo_aju.strip()
                nomes_existentes = [str(a.get("nome", "")).lower() for a in st.session_state["ajudantes"]]

                if novo_aju.lower() in nomes_existentes:
                    st.error("Este ajudante já está cadastrado.")
                else:
                    doc_id = f"aju_{int(agora_br().timestamp() * 1000)}"
                    dados_aju = {"id": doc_id, "nome": novo_aju}
                    sucesso = salvar_documento("ajudantes", doc_id, dados_aju)

                    if sucesso:
                        st.session_state["ajudantes"].append(dados_aju)
                        st.success(f"Ajudante {novo_aju} adicionado!")
                        st.rerun()

        st.markdown("---")
        st.write("**Ajudantes Atuais:**")

        if not st.session_state["ajudantes"]:
            st.info("Nenhum ajudante cadastrado.")

        for a_obj in st.session_state["ajudantes"]:
            a_nome = a_obj.get("nome", "")
            a_id = a_obj.get("id", a_nome)
            c_aju1, c_aju2 = st.columns([4, 2])

            c_aju1.write(f"👤 {a_nome}")

            if c_aju2.button("Excluir", key=f"del_aju_{a_id}"):
                st.session_state[f"confirmar_aju_{a_id}"] = True

            if st.session_state.get(f"confirmar_aju_{a_id}", False):
                st.warning(f"Excluir o ajudante '{a_nome}'?")
                ca1, ca2 = st.columns(2)

                with ca1:
                    if st.button("Confirmar", key=f"conf_aju_{a_id}"):
                        sucesso = deletar_documento("ajudantes", a_id)
                        if sucesso:
                            st.session_state["ajudantes"] = [a for a in st.session_state["ajudantes"] if a.get("id") != a_id]
                            st.rerun()

                with ca2:
                    if st.button("Cancelar", key=f"canc_aju_{a_id}"):
                        st.session_state[f"confirmar_aju_{a_id}"] = False
                        st.rerun()


    st.markdown("---")
    st.markdown("### 🚛 Frota / Veículos")
    with st.form("form_cad_veiculo", clear_on_submit=True):
        fv1, fv2, fv3 = st.columns(3)
        with fv1:
            nova_placa = st.text_input("Placa", placeholder="ABC1D23")
        with fv2:
            novo_modelo = st.text_input("Modelo", placeholder="Ex: Mercedes Atego 1719")
        with fv3:
            capacidade_kg = st.number_input("Capacidade (kg)", min_value=0, step=100)
        cad_veic_btn = st.form_submit_button("Cadastrar Veículo")
        if cad_veic_btn and nova_placa:
            placa = nova_placa.upper().replace("-", "").strip()
            existentes = [str(v.get("placa", "")).upper().replace("-", "") for v in st.session_state["veiculos"]]
            if placa in existentes:
                st.error("Este veículo já está cadastrado.")
            else:
                doc_id = f"veic_{placa}"
                dados = {"id": doc_id, "placa": placa, "modelo": novo_modelo.strip(), "capacidade_kg": int(capacidade_kg), "ativo": True}
                if salvar_documento("veiculos", doc_id, dados):
                    st.session_state["veiculos"].append(dados)
                    st.success(f"Veículo {placa} cadastrado.")
                    st.rerun()

    if st.session_state["veiculos"]:
        df_veiculos = pd.DataFrame(st.session_state["veiculos"])
        cols_v = [c for c in ["placa", "modelo", "capacidade_kg", "ativo"] if c in df_veiculos.columns]
        st.dataframe(df_veiculos[cols_v], use_container_width=True, hide_index=True)


# ============================================================
# 4. OCORRÊNCIAS
# ============================================================

elif menu == "🚨 Ocorrências":
    ids_cargas = [str(c.get("id")) for c in cargas_lista]
    if not ids_cargas:
        st.info("Cadastre uma carga antes de registrar ocorrências.")
    else:
        with st.form("form_ocorrencia", clear_on_submit=True):
            oc1, oc2 = st.columns(2)
            with oc1:
                carga_oc = st.selectbox("Planejamento / Carga", ids_cargas)
                tipo_oc = st.selectbox("Tipo", ["Atraso no carregamento", "Trânsito", "Pane mecânica", "Pneu", "Cliente fechado", "Endereço incorreto", "Falta de mercadoria", "Avaria", "Recusa", "Reentrega", "Documentação", "Outro"])
            with oc2:
                responsavel_oc = st.text_input("Responsável pelo tratamento")
                prioridade_oc = st.selectbox("Prioridade", ["Baixa", "Média", "Alta", "Crítica"], index=1)
            descricao_oc = st.text_area("Descrição da ocorrência")
            if st.form_submit_button("🚨 Abrir Ocorrência", type="primary"):
                oc_id = f"oc_{int(agora_br().timestamp()*1000)}"
                dados_oc = {"id": oc_id, "carga_id": carga_oc, "tipo": tipo_oc, "responsavel": responsavel_oc.strip(), "prioridade": prioridade_oc, "descricao": descricao_oc.strip(), "status": "Aberta", "aberta_em": iso_agora_br()}
                if salvar_documento("ocorrencias", oc_id, dados_oc):
                    st.session_state["ocorrencias"].append(dados_oc)
                    st.success("Ocorrência registrada.")
                    st.rerun()

    abertas = [o for o in st.session_state.get("ocorrencias", []) if o.get("status", "Aberta") == "Aberta"]
    st.markdown(f"### Abertas ({len(abertas)})")
    if not abertas:
        st.success("Nenhuma ocorrência aberta.")
    for oc in abertas:
        with st.container(border=True):
            st.markdown(f"**#{oc.get('carga_id')} · {oc.get('tipo')} · {oc.get('prioridade', 'Média')}**")
            st.write(oc.get("descricao", ""))
            st.caption(f"Responsável: {oc.get('responsavel') or 'Não definido'} · Aberta em: {oc.get('aberta_em', '')}")
            if st.button("✅ Resolver", key=f"resolver_{oc.get('id')}"):
                if atualizar_campos_documento("ocorrencias", oc.get("id"), {"status": "Resolvida", "resolvida_em": iso_agora_br()}):
                    oc["status"] = "Resolvida"
                    st.rerun()


# ============================================================
# 5. RELATÓRIOS
# ============================================================

elif menu == "📈 Relatórios":


    if not cargas_lista:
        st.info("Nenhuma carga cadastrada para gerar relatórios.")
    else:
        st.markdown("### 🔎 Filtros do Relatório")
        rf1, rf2, rf3 = st.columns(3)

        with rf1:
            filtro_motorista = st.selectbox("Motorista", ["Todos"] + motoristas_lista, key="rel_motorista")

        with rf2:
            filtro_status = st.selectbox(
                "Status",
                [
                    "Todos",
                    "Aguardando Carregamento",
                    "Carregado / No Pátio",
                    "Em Trânsito / Viagem Iniciada",
                    "Entregue / Concluído",
                ],
                key="rel_status"
            )

        with rf3:
            filtro_busca = st.text_input("Pesquisar ID ou destino", placeholder="Ex: 1042 ou Uberaba", key="rel_busca")

        cargas_relatorio = list(cargas_lista)

        if filtro_motorista != "Todos":
            cargas_relatorio = [c for c in cargas_relatorio if c.get("motorista") == filtro_motorista]

        if filtro_status != "Todos":
            cargas_relatorio = [c for c in cargas_relatorio if c.get("status") == filtro_status]

        if filtro_busca:
            termo = filtro_busca.lower().strip()
            cargas_relatorio = [
                c for c in cargas_relatorio
                if termo in str(c.get("id", "")).lower() or termo in str(c.get("destino", "")).lower()
            ]

        total_relatorio = len(cargas_relatorio)
        entregues_relatorio = sum(1 for c in cargas_relatorio if c.get("status") == "Entregue / Concluído")
        transito_relatorio = sum(1 for c in cargas_relatorio if c.get("status") == "Em Trânsito / Viagem Iniciada")
        atrasadas_relatorio = sum(1 for c in cargas_relatorio if carga_atrasada(c))
        percentual_entregues = ((entregues_relatorio / total_relatorio) * 100) if total_relatorio else 0
        otd_percentual, otd_amostra = calcular_otd(cargas_relatorio)

        r1, r2, r3, r4, r5 = st.columns(5)

        with r1:
            st.markdown(f"""
                <div class="metric-card metric-blue">
                    <div class="metric-title">CARGAS</div>
                    <div class="metric-value">{total_relatorio}</div>
                    <div class="metric-subtitle">Resultado do filtro</div>
                </div>
                """, unsafe_allow_html=True)

        with r2:
            st.markdown(f"""
                <div class="metric-card metric-green">
                    <div class="metric-title">ENTREGUES</div>
                    <div class="metric-value">{entregues_relatorio}</div>
                    <div class="metric-subtitle">Concluídas</div>
                </div>
                """, unsafe_allow_html=True)

        with r3:
            st.markdown(f"""
                <div class="metric-card metric-yellow">
                    <div class="metric-title">EM TRÂNSITO</div>
                    <div class="metric-value">{transito_relatorio}</div>
                    <div class="metric-subtitle">Viagens em andamento</div>
                </div>
                """, unsafe_allow_html=True)

        with r4:
            st.markdown(f"""
                <div class="metric-card metric-red">
                    <div class="metric-title">ATRASADAS</div>
                    <div class="metric-value">{atrasadas_relatorio}</div>
                    <div class="metric-subtitle">Fora do prazo</div>
                </div>
                """, unsafe_allow_html=True)

        with r5:
            st.markdown(f"""
                <div class="metric-card metric-purple">
                    <div class="metric-title">OTD</div>
                    <div class="metric-value">{otd_percentual:.1f}%</div>
                    <div class="metric-subtitle">{otd_amostra} entrega(s) com realizado</div>
                </div>
                """, unsafe_allow_html=True)

        st.markdown("---")
        st.markdown('<div class="section-title">📊 Visão Gerencial</div>', unsafe_allow_html=True)

        cores_status = {
            "Aguardando Carregamento": "#3b82f6",
            "Carregado / No Pátio": "#22c55e",
            "Em Trânsito / Viagem Iniciada": "#f59e0b",
            "Entregue / Concluído": "#a855f7",
            "Sem Status": "#64748b",
        }
        status_ordem = [
            "Aguardando Carregamento",
            "Carregado / No Pátio",
            "Em Trânsito / Viagem Iniciada",
            "Entregue / Concluído",
        ]

        status_contagem = {status: 0 for status in status_ordem}
        for c in cargas_relatorio:
            status = c.get("status", "Sem Status")
            status_contagem[status] = status_contagem.get(status, 0) + 1

        motorista_contagem = {}
        for c in cargas_relatorio:
            motorista = c.get("motorista", "Sem motorista") or "Sem motorista"
            motorista_contagem[motorista] = motorista_contagem.get(motorista, 0) + 1

        evolucao_contagem = {}
        for c in cargas_relatorio:
            data_obj = converter_para_data(c.get("data_saida") or c.get("data_carga"))
            if data_obj:
                evolucao_contagem[data_obj] = evolucao_contagem.get(data_obj, 0) + 1

        def grafico_status_html(contagem, total):
            partes = []
            inicio = 0.0
            for status in status_ordem:
                valor = contagem.get(status, 0)
                if not valor or not total:
                    continue
                fim = inicio + (valor / total) * 100
                partes.append(f"{cores_status[status]} {inicio:.2f}% {fim:.2f}%")
                inicio = fim
            if not partes:
                return '<div class="chart-empty">Nenhuma carga encontrada.</div>'
            gradient = ", ".join(partes)
            legenda = "".join(
                f'<div class="legend-row"><span class="legend-dot" style="background:{cores_status[s]}"></span><span>{s}</span><b>{contagem.get(s,0)}</b></div>'
                for s in status_ordem if contagem.get(s, 0)
            )
            return f"""
            <div class="chart-card">
              <div class="chart-heading"><span>Cargas por Status</span><small>Distribuição atual</small></div>
              <div class="donut-layout">
                <div class="donut" style="background:conic-gradient({gradient})">
                  <div class="donut-hole"><strong>{total}</strong><span>cargas</span></div>
                </div>
                <div class="legend-list">{legenda}</div>
              </div>
            </div>"""

        def grafico_barras_html(titulo, dados, limite=8):
            pares = sorted(dados.items(), key=lambda x: x[1], reverse=True)[:limite]
            if not pares:
                return f'<div class="chart-card"><div class="chart-heading"><span>{titulo}</span></div><div class="chart-empty">Nenhum dado disponível.</div></div>'
            maximo = max(v for _, v in pares) or 1
            linhas = ""
            for nome, valor in pares:
                largura = max(7, (valor / maximo) * 100)
                linhas += f"""
                <div class="bar-row">
                  <div class="bar-label" title="{nome}">{nome}</div>
                  <div class="bar-track"><div class="bar-fill" style="width:{largura:.1f}%"></div></div>
                  <div class="bar-value">{valor}</div>
                </div>"""
            return f"""
            <div class="chart-card">
              <div class="chart-heading"><span>{titulo}</span><small>Top {len(pares)}</small></div>
              <div class="bar-list">{linhas}</div>
            </div>"""

        st.markdown(grafico_status_html(status_contagem, total_relatorio), unsafe_allow_html=True)
        st.markdown(grafico_barras_html("Cargas por Motorista", motorista_contagem), unsafe_allow_html=True)

        if evolucao_contagem:
            datas = sorted(evolucao_contagem)
            max_val = max(evolucao_contagem.values()) or 1
            pontos = []
            for data_obj in datas:
                valor = evolucao_contagem[data_obj]
                altura = max(8, (valor / max_val) * 100)
                pontos.append(
                    f'<div class="evo-item"><div class="evo-value">{valor}</div><div class="evo-bar" style="height:{altura:.1f}%"></div><div class="evo-label">{data_obj.strftime("%d/%m")}</div></div>'
                )
            html_evo = f"""
            <div class="chart-card">
              <div class="chart-heading"><span>Movimentação por Data</span><small>Data de saída/carregamento</small></div>
              <div class="evo-chart">{"".join(pontos)}</div>
            </div>"""
            st.markdown(html_evo, unsafe_allow_html=True)

        st.markdown(grafico_barras_html("Top Motoristas", motorista_contagem, limite=5), unsafe_allow_html=True)

        st.markdown("### 📋 Detalhamento das Cargas")
        df_tabela = preparar_dataframe(cargas_relatorio)

        if df_tabela.empty:
            st.info("Nenhuma carga encontrada com os filtros selecionados.")
        else:
            st.dataframe(df_tabela, use_container_width=True, hide_index=True)

        st.markdown("### 📥 Exportar Arquivos")
        col_exp1, col_exp2 = st.columns(2)

        with col_exp1:
            excel_data = gerar_excel_profissional(df_tabela)
            st.download_button(
                label="📥 Baixar Planilha Excel (.xlsx)",
                data=excel_data,
                file_name="relatorio_de_cargas.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

        with col_exp2:
            pdf_bytes = gerar_pdf(df_tabela)
            st.download_button(
                label="📄 Baixar Relatório em PDF",
                data=pdf_bytes,
                file_name="relatorio_de_cargas.pdf",
                mime="application/pdf",
                use_container_width=True
            )


st.markdown(
    '<div class="app-footer">Sistema de Gestão de Cargas e Logística • Todos os direitos reservados</div>',
    unsafe_allow_html=True,
)
